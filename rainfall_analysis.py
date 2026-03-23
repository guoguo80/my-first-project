"""
降雨数据分析模块（增强版）
支持KTS-03-D202双列分析、KT-STORM缩放、清零处理、小时分段计算、日降雨量替代及一致性检查
"""

import pandas as pd
import os
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# 颜色定义
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def analyze_rainfall_batch(input_dir, output_file, progress_callback=None):
    """批量分析降雨数据，支持多设备类型"""
    results = []

    try:
        if not os.path.exists(input_dir):
            raise FileNotFoundError(f"输入目录不存在：{input_dir}")

        # 获取所有Excel文件
        all_files = [f for f in os.listdir(input_dir) if f.endswith(('.xlsx', '.xls'))]
        # 过滤：必须包含指定型号或“雨量计”，且不含“状态”
        excel_files = []
        for f in all_files:
            if "状态" in f:
                continue
            if ("KTS-03" in f) or ("KTS-03-D202" in f) or ("KT-STORM" in f) or ("雨量计" in f):
                excel_files.append(f)

        if not excel_files:
            msg = "输入目录下无符合降雨分析条件的文件（需含 KTS-03/KTS-03-D202/KT-STORM/雨量计 且无“状态”）"
            results.append({"status": "warning", "message": msg})
            return results

        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for file_idx, file_name in enumerate(excel_files, 1):
                file_path = os.path.join(input_dir, file_name)
                base_sheet = os.path.splitext(file_name)[0]

                msg = f"\n🔄 处理第{file_idx}/{len(excel_files)}个文件：{file_name}"
                results.append({"status": "info", "message": msg})
                if progress_callback:
                    progress_callback(msg)

                # 读取原始数据（共用）
                try:
                    df_raw = pd.read_excel(file_path, sheet_name="数据")
                except Exception as e:
                    msg = f"❌ 读取文件 {file_name} 失败：{e}"
                    results.append({"status": "error", "message": msg})
                    continue

                # 判断设备类型
                is_d202 = "KTS-03-D202" in file_name and "状态" not in file_name
                is_storm = "KT-STORM" in file_name and "状态" not in file_name

                # 定义处理任务列表
                tasks = []
                if is_d202:
                    # 需要两次分析
                    tasks.append(("降雨量", 1.0, base_sheet))
                    tasks.append(("修正雨量", 1.0, base_sheet + "修正"))
                elif is_storm:
                    tasks.append(("降雨量", 0.1, base_sheet))
                else:
                    # 其他设备（KTS-03、雨量计等）按原始列处理
                    tasks.append(("降雨量", 1.0, base_sheet))

                for rainfall_col, scale_factor, sheet_name in tasks:
                    # 检查列是否存在
                    if rainfall_col not in df_raw.columns:
                        msg = f"⚠️ 文件 {file_name} 缺少列 '{rainfall_col}'，跳过该任务"
                        results.append({"status": "warning", "message": msg})
                        continue

                    # 处理单个任务
                    process_single_file(df_raw, rainfall_col, scale_factor, sheet_name, writer, results, progress_callback)

        msg = f"\n🎉 所有文件处理完成！结果汇总至：{output_file}"
        results.append({"status": "success", "message": msg})
        if progress_callback:
            progress_callback(msg)

    except Exception as e:
        msg = f"\n❌ 批量分析失败：{str(e)}"
        results.append({"status": "error", "message": msg})
        if progress_callback:
            progress_callback(msg)

    return results


def process_single_file(df_raw, rainfall_col, scale_factor, sheet_name, writer, results, progress_callback):
    """处理单个数据列，写入指定工作表"""
    try:
        # 复制数据，避免修改原始
        df = df_raw.copy()

        # 预处理
        df['上传日期'] = pd.to_datetime(df['上传日期'], errors='coerce')
        df = df.dropna(subset=['上传日期', rainfall_col])
        df = df.drop_duplicates(subset=['上传日期']).sort_values('上传日期').reset_index(drop=True)

        if df.empty:
            msg = f"⚠️ {sheet_name}：无有效数据，跳过"
            results.append({"status": "warning", "message": msg})
            return

        # 标记异常值
        df['is_abnormal'] = df[rainfall_col].isin([9998, 9999])

        # 缩放（仅对正常值）
        if scale_factor != 1.0:
            df.loc[~df['is_abnormal'], rainfall_col] = df.loc[~df['is_abnormal'], rainfall_col] * scale_factor

        # 创建有效值列（异常置为NaN）
        df['value'] = df[rainfall_col].where(~df['is_abnormal'], np.nan)

        # 累计降雨总量（有效值）
        valid_df = df.dropna(subset=['value'])
        if len(valid_df) >= 2:
            total_rainfall = valid_df.iloc[-1]['value'] - valid_df.iloc[0]['value']
        else:
            total_rainfall = 0.0

        # 最大分钟降雨强度（相邻有效值且时间差60秒）
        valid_df['next_time'] = valid_df['上传日期'].shift(-1)
        valid_df['next_value'] = valid_df['value'].shift(-1)
        valid_df['time_diff'] = (valid_df['next_time'] - valid_df['上传日期']).dt.total_seconds()
        pairs = valid_df[(valid_df['time_diff'] == 60) & valid_df['next_value'].notna()].copy()
        pairs['intensity'] = pairs['next_value'] - pairs['value']
        positive_pairs = pairs[pairs['intensity'] > 0]
        if not positive_pairs.empty:
            max_minute_intensity = positive_pairs['intensity'].max()
            max_min_idx = positive_pairs['intensity'].idxmax()
            max_min_time = positive_pairs.loc[max_min_idx, '上传日期']
            max_minute_period = f"{max_min_time.strftime('%Y-%m-%d %H:%M')}-{(max_min_time + pd.Timedelta(minutes=1)).strftime('%H:%M')}"
        else:
            max_minute_intensity = 0.0
            max_minute_period = "-"

        # ---------- 小时降雨量计算 ----------
        # 生成所有整点时刻（从最早整点到最晚整点）
        start_hour = df['上传日期'].min().floor('h')
        end_hour = df['上传日期'].max().ceil('h')
        all_hours = pd.date_range(start=start_hour, end=end_hour, freq='h')

        # 构建整点字典
        df_time = df.set_index('上传日期')
        hourly_values = {}      # 有效值，NaN表示缺失或异常
        hourly_abnormal = {}    # 是否异常（仅当存在时）
        for h in all_hours:
            if h in df_time.index:
                val = df_time.loc[h, 'value']
                ab = df_time.loc[h, 'is_abnormal']
                hourly_values[h] = val if not ab else np.nan
                hourly_abnormal[h] = ab
            else:
                hourly_values[h] = np.nan
                hourly_abnormal[h] = False

        # 存储小时结果：每个小时 (date, hour, value_str, has_clear)
        hourly_results = []
        # 按日期存储小时值和清零标记
        date_hourly = {}          # {date: [24个值]}
        date_clear = {}           # {date: [24个布尔]}

        for i in range(len(all_hours) - 1):
            T = all_hours[i]
            next_T = all_hours[i + 1]
            date = T.date()
            hour = T.hour

            # 检查该小时内是否有原始数据（用于判断整点缺失时是否无数据）
            mask_in_hour = (df['上传日期'] >= T) & (df['上传日期'] < next_T)
            has_data_in_hour = mask_in_hour.any()

            # 整点状态
            val_T = hourly_values[T]
            val_next = hourly_values[next_T]
            ab_T = hourly_abnormal[T]
            ab_next = hourly_abnormal[next_T]
            exists_T = not pd.isna(val_T) or ab_T
            exists_next = not pd.isna(val_next) or ab_next

            # 初始化小时值和清零标记
            hour_value = None
            has_clear = False

            if not exists_T and not exists_next:
                # 两个整点都缺失
                if has_data_in_hour:
                    hour_value = "9998"
                else:
                    hour_value = "9997"
            elif not exists_T or not exists_next:
                # 缺少一个整点
                hour_value = "9998"
            elif ab_T or ab_next:
                # 有异常整点
                hour_value = "9999"
            else:
                # 两个整点都存在且正常，进行分段计算
                # 获取该小时内所有有效值点（包括T和next_T）
                mask_valid = (df['上传日期'] >= T) & (df['上传日期'] <= next_T) & df['value'].notna()
                points = df.loc[mask_valid, ['上传日期', 'value']].sort_values('上传日期')
                if len(points) < 2:
                    # 理论上不应发生，但以防万一
                    hour_value = "9998"
                else:
                    total = 0.0
                    current_start = points.iloc[0]['value']
                    prev_val = current_start
                    for idx in range(1, len(points)):
                        curr_val = points.iloc[idx]['value']
                        if curr_val < prev_val:  # 下降（清零）
                            total += prev_val - current_start
                            current_start = curr_val
                            has_clear = True
                        prev_val = curr_val
                    # 最后一段到最后一个点（即next_T）
                    total += prev_val - current_start
                    hour_value = round(total, 2)

            # 存储结果
            hourly_results.append((date, hour, hour_value, has_clear))

            # 更新按日期存储的字典
            if date not in date_hourly:
                date_hourly[date] = [None] * 24
                date_clear[date] = [False] * 24
            date_hourly[date][hour] = hour_value
            date_clear[date][hour] = has_clear

        # 获取所有有数据的日期
        all_dates = sorted(df['上传日期'].dt.date.unique())

        # ---------- 过滤最后一日仅含一条00:00数据 ----------
        if len(all_dates) > 0:
            last_date = all_dates[-1]
            last_day_data = df[df['上传日期'].dt.date == last_date]
            if len(last_day_data) == 1:
                first_time = last_day_data.iloc[0]['上传日期'].time()
                if first_time == pd.to_datetime('00:00:00').time():
                    all_dates = all_dates[:-1]  # 移除最后一日

        # 构建小时横向表（使用过滤后的 all_dates）
        hourly_rows = []
        for date in all_dates:
            row = {'日期': date}
            # 默认所有小时为9997（无数据）
            hours = ["9997"] * 24
            clears = [False] * 24
            if date in date_hourly:
                for h in range(24):
                    if date_hourly[date][h] is not None:
                        hours[h] = date_hourly[date][h]
                        clears[h] = date_clear[date][h]
            for h in range(24):
                row[f'{h}时'] = hours[h]
            hourly_rows.append(row)

        hourly_df = pd.DataFrame(hourly_rows)
        if not hourly_df.empty:
            hourly_df = hourly_df.sort_values('日期').reset_index(drop=True)

        # ---------- 日降雨量计算 ----------
        # 找出所有零点记录
        zero_mask = df['上传日期'].dt.time == pd.to_datetime('00:00:00').time()
        df_zero = df[zero_mask].copy()
        df_zero = df_zero.set_index('上传日期')
        zero_dict = {}
        for date in all_dates:
            zero_time = pd.Timestamp(date).replace(hour=0, minute=0, second=0)
            if zero_time in df_zero.index:
                val = df_zero.loc[zero_time, 'value']
                ab = df_zero.loc[zero_time, 'is_abnormal']
                if not ab and not pd.isna(val):
                    zero_dict[date] = val
                else:
                    zero_dict[date] = None
            else:
                zero_dict[date] = None

        # 计算日降雨量，并记录是否使用替代（黄色标记）
        daily_rain = {}
        daily_yellow = {}
        dates_sorted = all_dates  # 已经排序
        for i, date in enumerate(dates_sorted):
            next_date = dates_sorted[i + 1] if i + 1 < len(dates_sorted) else None
            z = zero_dict[date]
            z_next = zero_dict[next_date] if next_date else None

            # 当日有效值
            day_data = df[df['上传日期'].dt.date == date].dropna(subset=['value'])
            if day_data.empty:
                daily_rain[date] = None
                daily_yellow[date] = False
                continue
            first_val = day_data.iloc[0]['value']
            last_val = day_data.iloc[-1]['value']

            if z is not None and z_next is not None:
                daily = z_next - z
                yellow = False
            elif z is not None and z_next is None:
                daily = last_val - z
                yellow = True
            elif z is None and z_next is not None:
                daily = z_next - first_val
                yellow = True
            else:
                daily = last_val - first_val
                yellow = True
            daily_rain[date] = round(daily, 2) if daily is not None else None
            daily_yellow[date] = yellow

        # 将日降雨量合并到小时表
        hourly_df['日降雨量'] = hourly_df['日期'].map(lambda d: daily_rain.get(d, None))

        # ---------- 过滤日降雨量为0的行 ----------
        if not hourly_df.empty:
            hourly_df = hourly_df[hourly_df['日降雨量'] != 0]
            hourly_df = hourly_df.reset_index(drop=True)

        # ---------- 小时合计与日降雨量一致性检查 ----------
        def sum_hours(row):
            total = 0.0
            for h in range(24):
                val = row[f'{h}时']
                if isinstance(val, (int, float)):
                    total += val
            return round(total, 2)

        hourly_df['小时合计'] = hourly_df.apply(sum_hours, axis=1)
        hourly_df['match'] = hourly_df.apply(
            lambda row: abs(row['小时合计'] - row['日降雨量']) < 0.01 if pd.notna(row['日降雨量']) else False, axis=1
        )

        # 将红色掩码转换为列表，彻底避免索引错位
        red_mask_list = list(~hourly_df['match'])

        # ========== 调试输出 ==========
        print(f"\n[调试] Sheet: {sheet_name}")
        print(f"hourly_df 行数: {len(hourly_df)}")
        print(f"red_mask_list 长度: {len(red_mask_list)}")
        print("红色掩码列表 (True表示应填充红色):", red_mask_list)
        for i in range(len(hourly_df)):
            row = hourly_df.iloc[i]
            print(f"  索引 {i}: 日期={row['日期']}, 日降雨量={row['日降雨量']}, 小时合计={row['小时合计']}, 匹配={row['match']}, 红色标记={red_mask_list[i]}")
        # =============================

        # ---------- 最大小时降雨量（用于汇总） ----------
        max_hour_val = 0.0
        max_hour_time = None
        for date, hour, val_str, _ in hourly_results:
            if isinstance(val_str, (int, float)):
                if val_str > max_hour_val:
                    max_hour_val = val_str
                    max_hour_time = pd.Timestamp(date).replace(hour=hour, minute=0)
        max_hour_period = max_hour_time.strftime('%Y-%m-%d %H:%M') if max_hour_time else "-"

        # ---------- 最大日降雨量 ----------
        valid_daily = [v for v in daily_rain.values() if v is not None]
        max_daily = max(valid_daily) if valid_daily else 0.0
        max_daily_date = None
        if max_daily > 0:
            for d, v in daily_rain.items():
                if v == max_daily:
                    max_daily_date = d
                    break
        max_daily_date_str = str(max_daily_date) if max_daily_date else "-"

        # 降雨天数（日降雨量>0）
        rainy_days = sum(1 for v in daily_rain.values() if v is not None and v > 0)

        # ---------- 汇总表 ----------
        summary_stats = {
            "统计指标": [
                "统计区间",
                "累计降雨总量(mm)",
                "降雨天数",
                "最大日降雨量(mm)",
                "最大小时降雨量(mm)",
                "最大分钟降雨强度(mm/min)"
            ],
            "对应时间段": [
                f"{df['上传日期'].min().strftime('%Y-%m-%d %H:%M')}~{df['上传日期'].max().strftime('%Y-%m-%d %H:%M')}",
                "-",
                "-",
                max_daily_date_str,
                max_hour_period,
                max_minute_period
            ],
            "数值": [
                "-",
                round(total_rainfall, 2),
                rainy_days,
                round(max_daily, 2),
                round(max_hour_val, 2),
                round(max_minute_intensity, 2)
            ]
        }
        summary_df = pd.DataFrame(summary_stats)[['统计指标', '对应时间段', '数值']]

        # ---------- 写入Excel ----------
        summary_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index=False)
        if not hourly_df.empty:
            # 删除临时列
            hourly_out = hourly_df.drop(columns=['小时合计', 'match'])
            # 强制列顺序：日期、日降雨量、0~23时
            cols = ['日期', '日降雨量'] + [f'{h}时' for h in range(24)]
            hourly_out = hourly_out.reindex(columns=cols, fill_value=None)
            startrow = len(summary_df) + 2
            hourly_out.to_excel(writer, sheet_name=sheet_name, startrow=startrow, startcol=0, index=False)

        ws = writer.sheets[sheet_name]

        # 自动调整列宽
        for col in ws.iter_cols():
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # 应用颜色
        if not hourly_df.empty:
            # 修正：数据实际起始行号为 startrow + 2（因为表头占一行）
            data_start_row = startrow + 2
            print(f"[调试] startrow={startrow}, data_start_row={data_start_row}")  # 调试输出
            for i in range(len(hourly_df)):
                row_num = data_start_row + i

                # 红色行（小时合计与日降雨量不匹配）—— 使用列表按位置判断
                should_red = red_mask_list[i]
                print(f"  写入Excel 行 {row_num}: 日期={hourly_df.iloc[i]['日期']}, 红色标记={should_red}")

                if should_red:
                    ws.cell(row=row_num, column=1).fill = RED_FILL

                # 日降雨量黄色（替代计算）
                date = hourly_df.iloc[i]['日期']
                if daily_yellow.get(date, False):
                    cell = ws.cell(row=row_num, column=2)  # 日降雨量列
                    if cell.fill.fgColor.rgb != "FF0000":  # 不被红色覆盖
                        cell.fill = YELLOW_FILL

                # 小时单元格根据值着色
                for h in range(24):
                    col_num = 3 + h  # 第1列日期，第2列日降雨量，第3列开始是0时
                    cell = ws.cell(row=row_num, column=col_num)
                    val = cell.value
                    # 先检查是否已经有红色背景（整行红），若无则按规则设置
                    if cell.fill.fgColor.rgb == "FF0000":
                        continue
                    if val == "9997" or val == 9997:
                        cell.fill = RED_FILL
                    elif val in ("9998", "9999", 9998, 9999):
                        cell.fill = YELLOW_FILL
                    else:
                        # 检查该小时是否有清零
                        if date_clear.get(date, [False]*24)[h]:
                            cell.fill = RED_FILL

        msg = f"✅ {sheet_name} 分析完成"
        results.append({"status": "success", "message": msg})
        if progress_callback:
            progress_callback(msg)

    except Exception as e:
        msg = f"❌ {sheet_name} 处理失败：{str(e)}"
        results.append({"status": "error", "message": msg})
        if progress_callback:
            progress_callback(msg)