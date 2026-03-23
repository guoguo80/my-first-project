"""
设备偏差分析模块 - 完整版
包含所有原 deviation.txt 的功能
"""

import pandas as pd
import numpy as np
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings

warnings.filterwarnings('ignore')

# 设置matplotlib中文显示
try:
    plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
except:
    pass
plt.rcParams["axes.unicode_minus"] = False

# 雨量等级划分标准（小时降雨量，单位：mm）
RAINFALL_CLASSIFICATION = [
    ("小雨", 0, 2.5),
    ("中雨", 2.5, 8),
    ("大雨", 8, 16),
    ("暴雨", 16, 20),
    ("短时强降雨", 20, float('inf'))
]


def format_excel_worksheet(worksheet, df, start_row=1):
    """
    格式化Excel工作表：自动调整列宽，设置表头自动换行和样式
    """
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 处理表头（第一行）
    for col_num, column_title in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f"{col_letter}{start_row}"]

        # 设置表头样式
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 处理数据行
    for row_num, row_data in enumerate(df.values, start_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}{row_num}"]

            # 设置数据样式
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    # 自动调整列宽
    for col_num, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)

        # 计算表头长度
        max_length = len(str(column))

        # 计算该列数据最大长度
        for row_num in range(start_row + 1, len(df) + start_row + 1):
            cell_value = worksheet[f"{col_letter}{row_num}"].value
            if cell_value:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length

        # 设置列宽（限制最大宽度，避免过宽）
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    # 设置行高
    worksheet.row_dimensions[start_row].height = 40  # 表头行高
    for row in range(start_row + 1, len(df) + start_row + 1):
        worksheet.row_dimensions[row].height = 25  # 数据行高

    # 冻结表头行
    worksheet.freeze_panes = f"A{start_row + 1}"


def calculate_correlation_and_regression(x_data, y_data):
    """
    使用numpy计算相关系数和线性回归
    替代scipy.stats.linregress
    """
    # 移除NaN值
    mask = ~(np.isnan(x_data) | np.isnan(y_data))
    x_clean = x_data[mask]
    y_clean = y_data[mask]

    if len(x_clean) < 2:
        return None, None, None, None

    # 计算相关系数
    corr_matrix = np.corrcoef(x_clean, y_clean)
    r_value = corr_matrix[0, 1]

    # 计算线性回归参数（最小二乘法）
    n = len(x_clean)
    sum_x = np.sum(x_clean)
    sum_y = np.sum(y_clean)
    sum_xy = np.sum(x_clean * y_clean)
    sum_x2 = np.sum(x_clean ** 2)

    # 防止除零
    denominator = n * sum_x2 - sum_x ** 2
    if denominator == 0:
        slope = 0
    else:
        slope = (n * sum_xy - sum_x * sum_y) / denominator

    # 截距: (Σy - b*Σx) / n
    intercept = (sum_y - slope * sum_x) / n

    # 计算R²
    r_squared = r_value ** 2

    return slope, intercept, r_value, r_squared


def classify_rainfall_intensity(hourly_rainfall):
    """
    根据小时降雨量划分雨量等级
    """
    if pd.isna(hourly_rainfall) or hourly_rainfall <= 0:
        return "无降雨"
    elif hourly_rainfall <= 2.5:
        return "小雨"
    elif hourly_rainfall <= 8:
        return "中雨"
    elif hourly_rainfall <= 16:
        return "大雨"
    elif hourly_rainfall <= 20:
        return "暴雨"
    else:
        return "短时强降雨"


def analyze_accuracy_by_rainfall_class(data_dict, standard_device, output_dir):
    """
    新增函数：分析不同降雨等级下各设备相对于标准设备的计量准确度
    """
    results = []

    # 提取所有设备的小时数据
    hour_data_dict = {}
    rainfall_classes = ["小雨", "中雨", "大雨", "暴雨", "短时强降雨"]

    for device, data in data_dict.items():
        # 提取所有小时降雨数据并打平
        hour_columns = [f"{i}时" for i in range(24)]
        hour_data = []

        for idx, row in data.iterrows():
            for hour in range(24):
                hour_col = f"{hour}时"
                if hour_col in data.columns:
                    rainfall = row[hour_col]
                    if not pd.isna(rainfall) and rainfall > 0:
                        # 计算该小时的时间
                        hour_time = row["日期"] + pd.Timedelta(hours=hour)
                        hour_data.append({
                            "datetime": hour_time,
                            "hour": hour,
                            "rainfall": rainfall,
                            "class": classify_rainfall_intensity(rainfall)
                        })

        hour_data_dict[device] = pd.DataFrame(hour_data)
        msg = f"  设备 {device}: {len(hour_data)} 个有效小时数据点"
        results.append({"status": "info", "message": msg})

    # 检查标准设备是否存在
    if standard_device not in hour_data_dict:
        msg = f"❌ 错误: 指定的标准设备 '{standard_device}' 不存在!"
        results.append({"status": "error", "message": msg})
        return None, results

    # 将各设备数据与标准设备数据按时间对齐
    accuracy_results = {}

    for device in data_dict.keys():
        if device == standard_device:
            continue

        msg = f"\n  分析设备 {device} 相对于 {standard_device} 的准确度:"
        results.append({"status": "info", "message": msg})

        # 合并标准设备和当前设备的数据
        std_data = hour_data_dict[standard_device].copy()
        dev_data = hour_data_dict[device].copy()

        # 按时间合并
        merged_data = pd.merge(
            std_data,
            dev_data,
            on="datetime",
            suffixes=('_std', '_dev'),
            how='inner'
        )

        if len(merged_data) == 0:
            msg = f"    ⚠️ 无共同时间数据，跳过"
            results.append({"status": "warning", "message": msg})
            continue

        # 计算每个降雨等级的准确度指标
        device_results = {}

        for rain_class in rainfall_classes:
            # 筛选该降雨等级的数据
            class_data = merged_data[merged_data["class_std"] == rain_class].copy()

            if len(class_data) == 0:
                device_results[rain_class] = {
                    "样本数": 0,
                    "平均绝对误差(mm)": np.nan,
                    "平均相对误差(%)": np.nan,
                    "相关系数": np.nan,
                    "最大绝对误差(mm)": np.nan,
                    "误差标准差": np.nan
                }
                continue

            # 计算误差
            class_data["绝对误差"] = np.abs(class_data["rainfall_dev"] - class_data["rainfall_std"])
            class_data["相对误差(%)"] = np.abs(class_data["rainfall_dev"] - class_data["rainfall_std"]) / class_data[
                "rainfall_std"] * 100

            # 计算相关系数
            if len(class_data) >= 2:
                corr_matrix = np.corrcoef(class_data["rainfall_std"], class_data["rainfall_dev"])
                corr = corr_matrix[0, 1]
            else:
                corr = np.nan

            # 存储结果
            device_results[rain_class] = {
                "样本数": len(class_data),
                "平均绝对误差(mm)": round(class_data["绝对误差"].mean(), 3),
                "平均相对误差(%)": round(class_data["相对误差(%)"].mean(), 2),
                "相关系数": round(corr, 4) if not np.isnan(corr) else np.nan,
                "最大绝对误差(mm)": round(class_data["绝对误差"].max(), 3),
                "误差标准差": round(class_data["绝对误差"].std(), 3)
            }

            msg = f"    {rain_class}: {len(class_data)} 个样本, 平均绝对误差: {device_results[rain_class]['平均绝对误差(mm)']:.3f}mm"
            results.append({"status": "info", "message": msg})

        accuracy_results[device] = device_results

    return accuracy_results, results


def generate_accuracy_tables(accuracy_results, output_path, standard_device):
    """
    生成不同降雨等级下计量准确度分析表格
    """
    results = []

    try:
        # 如果文件不存在，先创建一个
        if not os.path.exists(output_path):
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                pass

        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            # 1. 各设备在不同降雨等级下的准确度汇总表
            summary_data = []
            rainfall_classes = ["小雨", "中雨", "大雨", "暴雨", "短时强降雨"]

            for device, results_dict in accuracy_results.items():
                for rain_class in rainfall_classes:
                    if rain_class in results_dict:
                        class_result = results_dict[rain_class]

                        summary_data.append({
                            "设备名称": device,
                            "降雨等级": rain_class,
                            "样本数": class_result["样本数"],
                            "平均绝对误差(mm)": class_result["平均绝对误差(mm)"],
                            "平均相对误差(%)": class_result["平均相对误差(%)"],
                            "相关系数": class_result["相关系数"],
                            "最大绝对误差(mm)": class_result["最大绝对误差(mm)"],
                            "误差标准差(mm)": class_result["误差标准差"]
                        })

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="降雨等级准确度汇总", index=False)

            # 2. 各降雨等级下设备准确度排名
            for rain_class in rainfall_classes:
                rank_data = []

                for device, results_dict in accuracy_results.items():
                    if rain_class in results_dict and results_dict[rain_class]["样本数"] > 0:
                        class_result = results_dict[rain_class]

                        rank_data.append({
                            "设备名称": device,
                            "样本数": class_result["样本数"],
                            "平均绝对误差(mm)": class_result["平均绝对误差(mm)"],
                            "平均相对误差(%)": class_result["平均相对误差(%)"],
                            "相关系数": class_result["相关系数"],
                            "最大绝对误差(mm)": class_result["最大绝对误差(mm)"],
                            "误差标准差(mm)": class_result["误差标准差"]
                        })

                if rank_data:
                    rank_df = pd.DataFrame(rank_data)
                    # 按平均绝对误差排序（误差越小越好）
                    rank_df = rank_df.sort_values("平均绝对误差(mm)")
                    rank_df.to_excel(writer, sheet_name=f"{rain_class}准确度排名", index=False)

            # 3. 各设备综合准确度评估
            overall_data = []

            for device, results_dict in accuracy_results.items():
                # 计算加权平均准确度（按样本数加权）
                total_samples = 0
                weighted_mae = 0
                weighted_mre = 0

                for rain_class in rainfall_classes:
                    if rain_class in results_dict and results_dict[rain_class]["样本数"] > 0:
                        samples = results_dict[rain_class]["样本数"]
                        total_samples += samples
                        weighted_mae += results_dict[rain_class]["平均绝对误差(mm)"] * samples
                        weighted_mre += results_dict[rain_class]["平均相对误差(%)"] * samples

                if total_samples > 0:
                    overall_mae = weighted_mae / total_samples
                    overall_mre = weighted_mre / total_samples
                else:
                    overall_mae = np.nan
                    overall_mre = np.nan

                # 评估等级
                if overall_mae <= 0.5:
                    accuracy_level = "优秀"
                elif overall_mae <= 1.0:
                    accuracy_level = "良好"
                elif overall_mae <= 2.0:
                    accuracy_level = "一般"
                else:
                    accuracy_level = "较差"

                overall_data.append({
                    "设备名称": device,
                    "总样本数": total_samples,
                    "加权平均绝对误差(mm)": round(overall_mae, 3) if not np.isnan(overall_mae) else np.nan,
                    "加权平均相对误差(%)": round(overall_mre, 2) if not np.isnan(overall_mre) else np.nan,
                    "准确度等级": accuracy_level
                })

            overall_df = pd.DataFrame(overall_data)
            overall_df = overall_df.sort_values("加权平均绝对误差(mm)")
            overall_df.to_excel(writer, sheet_name="综合准确度评估", index=False)

        # 格式化Excel文件
        wb = load_workbook(output_path)

        # 格式化每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 读取工作表数据到DataFrame
            if ws.max_row > 1 and ws.max_column > 1:
                data = []
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                    data.append(row)

                # 第一行为表头
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])

                    # 清除工作表内容（保留格式设置）
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                        for cell in row:
                            cell.value = None

                    # 重新写入数据
                    for col_idx, col_name in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)

                    for row_idx, row_data in enumerate(df.values, 2):
                        for col_idx, cell_value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)

                    # 应用格式化
                    format_excel_worksheet(ws, df)

        # 保存工作簿
        wb.save(output_path)

        msg = f"✅ 降雨等级准确度分析表已保存到Excel文件"
        results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"❌ 生成准确度表格失败: {str(e)}"
        results.append({"status": "error", "message": msg})

    return results


def analyze_rainfall_classification(data_dict, output_dir):
    """
    分析各设备小时降雨量等级分布
    """
    results = []

    classification_results = {}

    for device, data in data_dict.items():
        msg = f"\n分析设备: {device}"
        results.append({"status": "info", "message": msg})

        # 提取所有小时降雨数据
        hour_columns = [f"{i}时" for i in range(24)]
        hour_data_list = []

        for col in hour_columns:
            if col in data.columns:
                hour_data_list.append(data[col].values)

        if hour_data_list:
            # 合并所有小时数据
            all_hour_data = np.concatenate(hour_data_list)
            # 去除NaN值
            all_hour_data = all_hour_data[~np.isnan(all_hour_data)]

            # 统计雨量等级
            classification_counts = {
                "小雨": 0,
                "中雨": 0,
                "大雨": 0,
                "暴雨": 0,
                "短时强降雨": 0,
                "总小时数": len(all_hour_data)
            }

            classification_rainfall = {
                "小雨": 0.0,
                "中雨": 0.0,
                "大雨": 0.0,
                "暴雨": 0.0,
                "短时强降雨": 0.0,
                "总降雨量": 0.0
            }

            for rainfall in all_hour_data:
                rainfall_class = classify_rainfall_intensity(rainfall)
                if rainfall_class in classification_counts:
                    classification_counts[rainfall_class] += 1
                    classification_rainfall[rainfall_class] += rainfall
                    classification_rainfall["总降雨量"] += rainfall

            # 计算百分比
            classification_percentages = {}
            for class_name, count in classification_counts.items():
                if class_name != "总小时数":
                    percentage = (count / classification_counts["总小时数"]) * 100 if classification_counts[
                                                                                          "总小时数"] > 0 else 0
                    classification_percentages[f"{class_name}_占比(%)"] = round(percentage, 2)

            # 存储结果
            classification_results[device] = {
                "counts": classification_counts,
                "rainfall": classification_rainfall,
                "percentages": classification_percentages
            }

            msg = f"  总小时数: {classification_counts['总小时数']}"
            results.append({"status": "info", "message": msg})
            msg = f"  小雨: {classification_counts['小雨']} 小时 ({classification_percentages.get('小雨_占比(%)', 0)}%)"
            results.append({"status": "info", "message": msg})
            msg = f"  中雨: {classification_counts['中雨']} 小时 ({classification_percentages.get('中雨_占比(%)', 0)}%)"
            results.append({"status": "info", "message": msg})
            msg = f"  大雨: {classification_counts['大雨']} 小时 ({classification_percentages.get('大雨_占比(%)', 0)}%)"
            results.append({"status": "info", "message": msg})
            msg = f"  暴雨: {classification_counts['暴雨']} 小时 ({classification_percentages.get('暴雨_占比(%)', 0)}%)"
            results.append({"status": "info", "message": msg})
            msg = f"  短时强降雨: {classification_counts['短时强降雨']} 小时 ({classification_percentages.get('短时强降雨_占比(%)', 0)}%)"
            results.append({"status": "info", "message": msg})

    return classification_results, results


def generate_classification_tables(classification_results, output_path):
    """
    生成雨量等级分类统计表格
    """
    results = []

    try:
        # 如果文件不存在，先创建一个
        if not os.path.exists(output_path):
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                pass

        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            # 1. 各设备雨量等级小时数统计表
            hour_counts_data = []
            for device, results_dict in classification_results.items():
                counts = results_dict["counts"]
                percentages = results_dict["percentages"]

                hour_counts_data.append({
                    "设备名称": device,
                    "总小时数": counts["总小时数"],
                    "小雨(小时)": counts["小雨"],
                    "小雨占比(%)": percentages.get("小雨_占比(%)", 0),
                    "中雨(小时)": counts["中雨"],
                    "中雨占比(%)": percentages.get("中雨_占比(%)", 0),
                    "大雨(小时)": counts["大雨"],
                    "大雨占比(%)": percentages.get("大雨_占比(%)", 0),
                    "暴雨(小时)": counts["暴雨"],
                    "暴雨占比(%)": percentages.get("暴雨_占比(%)", 0),
                    "短时强降雨(小时)": counts["短时强降雨"],
                    "短时强降雨占比(%)": percentages.get("短时强降雨_占比(%)", 0)
                })

            hour_counts_df = pd.DataFrame(hour_counts_data)
            hour_counts_df.to_excel(writer, sheet_name="雨量等级小时数统计", index=False)

            # 2. 各设备雨量等级累计降雨量统计表
            rainfall_data = []
            for device, results_dict in classification_results.items():
                rainfall = results_dict["rainfall"]
                total_rainfall = rainfall["总降雨量"]

                rainfall_data.append({
                    "设备名称": device,
                    "总降雨量(mm)": round(total_rainfall, 2),
                    "小雨降雨量(mm)": round(rainfall["小雨"], 2),
                    "小雨占比(%)": round((rainfall["小雨"] / total_rainfall) * 100, 2) if total_rainfall > 0 else 0,
                    "中雨降雨量(mm)": round(rainfall["中雨"], 2),
                    "中雨占比(%)": round((rainfall["中雨"] / total_rainfall) * 100, 2) if total_rainfall > 0 else 0,
                    "大雨降雨量(mm)": round(rainfall["大雨"], 2),
                    "大雨占比(%)": round((rainfall["大雨"] / total_rainfall) * 100, 2) if total_rainfall > 0 else 0,
                    "暴雨降雨量(mm)": round(rainfall["暴雨"], 2),
                    "暴雨占比(%)": round((rainfall["暴雨"] / total_rainfall) * 100, 2) if total_rainfall > 0 else 0,
                    "短时强降雨量(mm)": round(rainfall["短时强降雨"], 2),
                    "短时强降雨占比(%)": round((rainfall["短时强降雨"] / total_rainfall) * 100,
                                               2) if total_rainfall > 0 else 0
                })

            rainfall_df = pd.DataFrame(rainfall_data)
            rainfall_df.to_excel(writer, sheet_name="雨量等级降雨量统计", index=False)

        # 格式化Excel文件
        wb = load_workbook(output_path)

        # 格式化每个工作表
        for sheet_name in ["雨量等级小时数统计", "雨量等级降雨量统计"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 读取工作表数据到DataFrame
                if ws.max_row > 1 and ws.max_column > 1:
                    data = []
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                        data.append(row)

                    # 第一行为表头
                    if data:
                        df = pd.DataFrame(data[1:], columns=data[0])

                        # 清除工作表内容（保留格式设置）
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                            for cell in row:
                                cell.value = None

                        # 重新写入数据
                        for col_idx, col_name in enumerate(df.columns, 1):
                            ws.cell(row=1, column=col_idx, value=col_name)

                        for row_idx, row_data in enumerate(df.values, 2):
                            for col_idx, cell_value in enumerate(row_data, 1):
                                ws.cell(row=row_idx, column=col_idx, value=cell_value)

                        # 应用格式化
                        format_excel_worksheet(ws, df)

        # 保存工作簿
        wb.save(output_path)

        msg = f"✅ 雨量等级分类统计表已保存到Excel文件"
        results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"❌ 生成分类表格失败: {str(e)}"
        results.append({"status": "error", "message": msg})

    return results


def generate_classification_charts(classification_results, standard_device, output_dir):
    """
    生成雨量等级分类对比图表
    """
    results = []

    try:
        # 提取设备列表
        devices = list(classification_results.keys())

        # 1. 各设备雨量等级小时数对比图（堆积柱状图）
        plt.figure(figsize=(14, 8))

        # 准备数据
        rainfall_classes = ["小雨", "中雨", "大雨", "暴雨", "短时强降雨"]
        colors = ['lightblue', 'green', 'orange', 'red', 'purple']

        # 创建底部位置
        bottom = np.zeros(len(devices))

        for idx, rainfall_class in enumerate(rainfall_classes):
            class_counts = [classification_results[device]["counts"][rainfall_class] for device in devices]

            bars = plt.bar(devices, class_counts, bottom=bottom, color=colors[idx],
                           edgecolor='black', label=rainfall_class, alpha=0.8)
            bottom += class_counts

        plt.title("各设备雨量等级小时数对比", fontsize=16, pad=20)
        plt.xlabel("设备名称", fontsize=14)
        plt.ylabel("小时数", fontsize=14)
        plt.xticks(rotation=45)
        plt.legend(title="雨量等级", loc="upper right")
        plt.grid(True, alpha=0.3, axis='y')

        # 添加总数标注
        total_counts = [classification_results[device]["counts"]["总小时数"] for device in devices]
        for i, total in enumerate(total_counts):
            plt.text(i, total + max(total_counts) * 0.01, f'{total}',
                     ha='center', va='bottom', fontsize=10)

        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "各设备雨量等级小时数对比.png"), dpi=150, bbox_inches='tight')
        plt.close()

        # 2. 各设备雨量等级小时数百分比对比图（堆积百分比柱状图）
        plt.figure(figsize=(14, 8))

        bottom = np.zeros(len(devices))

        for idx, rainfall_class in enumerate(rainfall_classes):
            class_percentages = []
            for device in devices:
                total_hours = classification_results[device]["counts"]["总小时数"]
                class_hours = classification_results[device]["counts"][rainfall_class]
                percentage = (class_hours / total_hours) * 100 if total_hours > 0 else 0
                class_percentages.append(percentage)

            bars = plt.bar(devices, class_percentages, bottom=bottom, color=colors[idx],
                           edgecolor='black', label=rainfall_class, alpha=0.8)
            bottom += class_percentages

            # 添加百分比标注
            for i, (bar, percentage) in enumerate(zip(bars, class_percentages)):
                if percentage > 5:  # 只显示较大百分比的标注
                    height = bar.get_height()
                    y_pos = bottom[i] - height / 2
                    plt.text(bar.get_x() + bar.get_width() / 2, y_pos,
                             f'{percentage:.1f}%', ha='center', va='center',
                             fontsize=9, color='white', fontweight='bold')

        plt.title("各设备雨量等级小时数百分比对比", fontsize=16, pad=20)
        plt.xlabel("设备名称", fontsize=14)
        plt.ylabel("百分比(%)", fontsize=14)
        plt.xticks(rotation=45)
        plt.legend(title="雨量等级", loc="upper right")
        plt.grid(True, alpha=0.3, axis='y')
        plt.ylim(0, 100)

        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "各设备雨量等级小时数百分比对比.png"), dpi=150, bbox_inches='tight')
        plt.close()

        # 3. 各设备雨量等级累计降雨量对比图
        plt.figure(figsize=(14, 8))

        bottom = np.zeros(len(devices))

        for idx, rainfall_class in enumerate(rainfall_classes):
            class_rainfall = [classification_results[device]["rainfall"][rainfall_class] for device in devices]

            bars = plt.bar(devices, class_rainfall, bottom=bottom, color=colors[idx],
                           edgecolor='black', label=rainfall_class, alpha=0.8)
            bottom += class_rainfall

        plt.title("各设备雨量等级累计降雨量对比", fontsize=16, pad=20)
        plt.xlabel("设备名称", fontsize=14)
        plt.ylabel("累计降雨量(mm)", fontsize=14)
        plt.xticks(rotation=45)
        plt.legend(title="雨量等级", loc="upper right")
        plt.grid(True, alpha=0.3, axis='y')

        # 添加总数标注
        total_rainfall = [classification_results[device]["rainfall"]["总降雨量"] for device in devices]
        for i, total in enumerate(total_rainfall):
            plt.text(i, total + max(total_rainfall) * 0.01, f'{total:.1f}',
                     ha='center', va='bottom', fontsize=10)

        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "各设备雨量等级累计降雨量对比.png"), dpi=150, bbox_inches='tight')
        plt.close()

        # 4. 各设备不同雨量等级的小时数雷达图
        fig = plt.figure(figsize=(10, 8))
        ax = fig.add_subplot(111, projection='polar')

        # 准备数据
        angles = np.linspace(0, 2 * np.pi, len(rainfall_classes), endpoint=False).tolist()
        angles += angles[:1]  # 闭合雷达图

        for device in devices:
            values = []
            for rainfall_class in rainfall_classes:
                total_hours = classification_results[device]["counts"]["总小时数"]
                class_hours = classification_results[device]["counts"][rainfall_class]
                # 标准化到0-100范围
                normalized_value = (class_hours / total_hours) * 100 if total_hours > 0 else 0
                values.append(normalized_value)

            values += values[:1]  # 闭合雷达图

            # 绘制雷达图
            ax.plot(angles, values, 'o-', linewidth=2, label=device, alpha=0.7)
            ax.fill(angles, values, alpha=0.1)

        # 设置雷达图标签
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(rainfall_classes, fontsize=12)
        ax.set_ylim(0, 100)
        ax.set_yticks([0, 25, 50, 75, 100])
        ax.set_yticklabels(['0%', '25%', '50%', '75%', '100%'], fontsize=10)
        ax.grid(True)

        plt.title("各设备雨量等级小时数分布雷达图", fontsize=16, pad=20)
        plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.0))
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "各设备雨量等级小时数分布雷达图.png"), dpi=150, bbox_inches='tight')
        plt.close()

        msg = f"✅ 雨量等级分类对比图表已保存到: {output_dir}"
        results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"⚠️ 生成分类图表时出错: {str(e)}"
        results.append({"status": "warning", "message": msg})

    return results


def read_rainfall_data(file_path):
    """读取Excel文件中的降雨数据"""
    results = []

    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"输入文件不存在：{file_path}")

        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names

        data_dict = {}

        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

                # 查找数据开始位置
                start_row = None
                for i in range(len(df)):
                    cell_value = df.iloc[i, 0]
                    if isinstance(cell_value, str) and "日期" in cell_value:
                        start_row = i
                        break

                if start_row is None:
                    msg = f"警告: 工作表 '{sheet_name}' 中未找到'日期'列，跳过该设备"
                    results.append({"status": "warning", "message": msg})
                    continue

                # 提取数据部分
                data = df.iloc[start_row:].reset_index(drop=True)
                data.columns = data.iloc[0]
                data = data.iloc[1:].reset_index(drop=True)

                # 清理数据
                data["日期"] = pd.to_datetime(data["日期"], errors="coerce")

                # 转换数值列
                numeric_cols = ["日降雨量"] + [f"{i}时" for i in range(24)]
                for col in numeric_cols:
                    if col in data.columns:
                        data[col] = pd.to_numeric(data[col], errors="coerce")
                        data[col] = data[col].replace([9999, 9998], np.nan)

                # 删除全为NaN的行
                data = data.dropna(subset=["日期"])
                data = data[~data["日降雨量"].isna()]

                data_dict[sheet_name] = data

                msg = f"✅ 成功读取设备 '{sheet_name}' 的数据: {len(data)} 天"
                results.append({"status": "success", "message": msg})

            except Exception as e:
                msg = f"❌ 读取设备 '{sheet_name}' 数据失败: {str(e)}"
                results.append({"status": "error", "message": msg})

        return data_dict, results

    except Exception as e:
        msg = f"❌ 读取降雨数据失败: {str(e)}"
        results.append({"status": "error", "message": msg})
        return None, results


def generate_accuracy_charts(accuracy_results, standard_device, output_dir):
    """
    生成不同降雨等级下计量准确度对比图表
    """
    results = []

    if not accuracy_results:
        msg = "⚠️ 无准确度分析结果，跳过图表生成"
        results.append({"status": "warning", "message": msg})
        return results

    try:
        rainfall_classes = ["小雨", "中雨", "大雨", "暴雨", "短时强降雨"]
        devices = list(accuracy_results.keys())

        # 1. 各设备在不同降雨等级下的平均绝对误差热力图
        plt.figure(figsize=(12, 8))

        # 准备数据矩阵
        data_matrix = []
        device_labels = []

        for device in devices:
            device_errors = []
            for rain_class in rainfall_classes:
                if rain_class in accuracy_results[device] and accuracy_results[device][rain_class]["样本数"] > 0:
                    device_errors.append(accuracy_results[device][rain_class]["平均绝对误差(mm)"])
                else:
                    device_errors.append(np.nan)

            # 只添加有数据的设备
            if not all(np.isnan(err) for err in device_errors):
                data_matrix.append(device_errors)
                device_labels.append(device)

        if data_matrix:
            data_matrix = np.array(data_matrix)

            # 创建热力图
            im = plt.imshow(data_matrix, cmap='RdYlGn_r', aspect='auto', vmin=0, vmax=3)

            # 添加颜色条
            plt.colorbar(im, label='平均绝对误差(mm)')

            # 设置坐标轴
            plt.xticks(range(len(rainfall_classes)), rainfall_classes, rotation=0)
            plt.yticks(range(len(device_labels)), device_labels)

            # 添加数值标签
            for i in range(len(device_labels)):
                for j in range(len(rainfall_classes)):
                    if not np.isnan(data_matrix[i, j]):
                        plt.text(j, i, f'{data_matrix[i, j]:.2f}',
                                 ha='center', va='center', color='black', fontsize=9)

            plt.title(f"各设备在不同降雨等级下的平均绝对误差(mm)\n(相对于标准设备: {standard_device})", fontsize=14,
                      pad=20)
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, "降雨等级准确度热力图.png"), dpi=150, bbox_inches='tight')
            plt.close()

        # 2. 各降雨等级下设备平均绝对误差对比图
        plt.figure(figsize=(14, 8))

        # 设置颜色
        colors = plt.cm.Set3(np.linspace(0, 1, len(rainfall_classes)))

        x = np.arange(len(devices))
        width = 0.15  # 柱状图宽度
        offsets = np.linspace(-width * len(rainfall_classes) / 2, width * len(rainfall_classes) / 2,
                              len(rainfall_classes))

        for idx, rain_class in enumerate(rainfall_classes):
            class_errors = []
            for device in devices:
                if rain_class in accuracy_results[device] and accuracy_results[device][rain_class]["样本数"] > 0:
                    class_errors.append(accuracy_results[device][rain_class]["平均绝对误差(mm)"])
                else:
                    class_errors.append(np.nan)

            # 绘制柱状图
            bars = plt.bar(x + offsets[idx], class_errors, width,
                           label=rain_class, color=colors[idx], alpha=0.8, edgecolor='black')

            # 添加数值标签
            for bar_idx, (bar, error) in enumerate(zip(bars, class_errors)):
                if not np.isnan(error) and error > 0:
                    height = bar.get_height()
                    plt.text(bar.get_x() + bar.get_width() / 2., height + 0.05,
                             f'{error:.2f}', ha='center', va='bottom', fontsize=8)

        plt.xlabel('设备名称', fontsize=12)
        plt.ylabel('平均绝对误差(mm)', fontsize=12)
        plt.title(f'各设备在不同降雨等级下的平均绝对误差对比\n(相对于标准设备: {standard_device})', fontsize=14, pad=20)
        plt.xticks(x, devices, rotation=45)
        plt.legend(title='降雨等级', loc='upper right')
        plt.grid(True, alpha=0.3, axis='y')
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "降雨等级准确度对比图.png"), dpi=150, bbox_inches='tight')
        plt.close()

        # 3. 各设备综合准确度雷达图
        plt.figure(figsize=(10, 8))
        ax = plt.subplot(111, projection='polar')

        # 准备数据
        angles = np.linspace(0, 2 * np.pi, len(rainfall_classes), endpoint=False).tolist()
        angles += angles[:1]  # 闭合雷达图

        for device in devices:
            values = []
            for rain_class in rainfall_classes:
                if rain_class in accuracy_results[device] and accuracy_results[device][rain_class]["样本数"] > 0:
                    # 使用误差的倒数作为准确度指标（误差越小，准确度越高）
                    error = accuracy_results[device][rain_class]["平均绝对误差(mm)"]
                    if error > 0:
                        # 归一化到0-100范围（误差1mm对应80分）
                        accuracy_score = max(0, 100 - error * 20)
                    else:
                        accuracy_score = 0
                else:
                    accuracy_score = 0
                values.append(accuracy_score)

            values += values[:1]  # 闭合雷达图

            # 绘制雷达图
            ax.plot(angles, values, 'o-', linewidth=2, label=device, alpha=0.7)
            ax.fill(angles, values, alpha=0.1)

        # 设置雷达图标签
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(rainfall_classes, fontsize=11)
        ax.set_ylim(0, 100)
        ax.set_yticks([0, 20, 40, 60, 80, 100])
        ax.set_yticklabels(['0', '20', '40', '60', '80', '100'], fontsize=9)
        ax.grid(True)

        plt.title(f"各设备在不同降雨等级下的准确度雷达图\n(分数越高表示准确度越高)", fontsize=14, pad=20)
        plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.0))
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "降雨等级准确度雷达图.png"), dpi=150, bbox_inches='tight')
        plt.close()

        # 4. 样本数量分布图
        plt.figure(figsize=(12, 6))

        sample_data = []
        for device in devices:
            for rain_class in rainfall_classes:
                if rain_class in accuracy_results[device]:
                    samples = accuracy_results[device][rain_class]["样本数"]
                    if samples > 0:
                        sample_data.append({
                            "设备": device,
                            "降雨等级": rain_class,
                            "样本数": samples
                        })

        if sample_data:
            sample_df = pd.DataFrame(sample_data)

            # 创建分组柱状图
            unique_devices = sample_df["设备"].unique()
            unique_classes = sample_df["降雨等级"].unique()

            x = np.arange(len(unique_devices))
            width = 0.15

            for idx, rain_class in enumerate(unique_classes):
                class_samples = []
                for device in unique_devices:
                    mask = (sample_df["设备"] == device) & (sample_df["降雨等级"] == rain_class)
                    if mask.any():
                        class_samples.append(sample_df.loc[mask, "样本数"].values[0])
                    else:
                        class_samples.append(0)

                offset = width * (idx - len(unique_classes) / 2 + 0.5)
                plt.bar(x + offset, class_samples, width, label=rain_class, alpha=0.8)

            plt.xlabel('设备名称', fontsize=12)
            plt.ylabel('样本数量', fontsize=12)
            plt.title('各降雨等级下的样本数量分布', fontsize=14, pad=20)
            plt.xticks(x, unique_devices, rotation=45)
            plt.legend(title='降雨等级')
            plt.grid(True, alpha=0.3, axis='y')
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, "降雨等级样本分布图.png"), dpi=150, bbox_inches='tight')
            plt.close()

        msg = f"✅ 降雨等级准确度对比图表已保存到: {output_dir}"
        results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"⚠️ 生成准确度图表时出错: {str(e)}"
        results.append({"status": "warning", "message": msg})

    return results


def generate_visualizations(data_dict, common_data, standard_device, output_dir, common_dates):
    """
    生成可视化图表
    """
    results = []

    try:
        # 1. 各设备累计降雨量对比图
        plt.figure(figsize=(12, 6))
        devices = list(data_dict.keys())
        total_rainfalls = [data_dict[dev]["日降雨量"].sum() for dev in devices]

        # 颜色映射：标准设备用特殊颜色
        colors = ['lightblue' if dev != standard_device else 'orange' for dev in devices]

        bars = plt.bar(devices, total_rainfalls, color=colors, edgecolor='black')
        plt.title("各设备累计降雨量对比", fontsize=16, pad=20)
        plt.xlabel("设备名称", fontsize=14)
        plt.ylabel("累计降雨量(mm)", fontsize=14)
        plt.xticks(rotation=45)

        # 在柱状图上显示数值
        for bar, value in zip(bars, total_rainfalls):
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2., height + max(total_rainfalls) * 0.01,
                     f'{value:.1f}', ha='center', va='bottom', fontsize=10)

        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "累计降雨量对比.png"), dpi=150, bbox_inches='tight')
        plt.close()

        msg = f"✅ 累计降雨量对比图已保存"
        results.append({"status": "success", "message": msg})

        # 2. 各设备日降雨量对比趋势图（共有日期）
        if common_dates and len(common_dates) > 0:
            plt.figure(figsize=(15, 8))

            # 定义不同颜色和标记样式
            colors = ['blue', 'green', 'red', 'purple', 'orange', 'brown', 'pink', 'gray']
            markers = ['o', 's', '^', 'v', '<', '>', 'd', 'p']

            for idx, device in enumerate(data_dict.keys()):
                if device in common_data and len(common_data[device]) > 0:
                    # 按日期排序
                    sorted_data = common_data[device].sort_values("日期")
                    color_idx = idx % len(colors)
                    marker_idx = idx % len(markers)

                    plt.plot(sorted_data["日期"], sorted_data["日降雨量"],
                             linewidth=2, marker=markers[marker_idx], markersize=4,
                             color=colors[color_idx], label=device)

            plt.title("各设备日降雨量对比趋势（共有日期）", fontsize=16, pad=20)
            plt.xlabel("日期", fontsize=14)
            plt.ylabel("日降雨量(mm)", fontsize=14)
            plt.xticks(rotation=45)
            plt.grid(True, alpha=0.3)
            plt.legend(loc="upper right", fontsize=10)
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, "日降雨量对比趋势.png"), dpi=150, bbox_inches='tight')
            plt.close()

            msg = f"✅ 日降雨量对比趋势图已保存"
            results.append({"status": "success", "message": msg})

        # 3. 与标准设备的差异散点图
        if common_dates and len(common_dates) > 0 and standard_device in common_data:
            # 计算需要多少个子图
            other_devices = [dev for dev in data_dict.keys() if dev != standard_device]
            num_other_devices = len(other_devices)

            if num_other_devices > 0:
                # 确定子图布局
                if num_other_devices <= 4:
                    rows, cols = 1, num_other_devices
                    figsize = (5 * cols, 5)
                elif num_other_devices <= 6:
                    rows, cols = 2, 3
                    figsize = (15, 10)
                else:
                    rows, cols = 2, 4
                    figsize = (20, 10)

                fig, axes = plt.subplots(rows, cols, figsize=figsize)

                # 如果只有一个子图，确保axes是数组
                if rows == 1 and cols == 1:
                    axes = [axes]
                elif rows == 1:
                    axes = axes.flatten()
                elif cols == 1:
                    axes = axes.flatten()
                else:
                    axes = axes.flatten()

                # 定义颜色
                colors = ['blue', 'green', 'red', 'purple', 'orange', 'brown', 'pink', 'gray']

                for idx, device in enumerate(other_devices):
                    if idx >= rows * cols:
                        break

                    if device in common_data:
                        ax = axes[idx]

                        # 获取共有日期的数据
                        merged_data = pd.merge(
                            common_data[standard_device][["日期", "日降雨量"]],
                            common_data[device][["日期", "日降雨量"]],
                            on="日期",
                            suffixes=('_标准', '_对比')
                        )

                        x_data = merged_data["日降雨量_标准"]
                        y_data = merged_data["日降雨量_对比"]

                        # 绘制散点图
                        color_idx = idx % len(colors)
                        ax.scatter(x_data, y_data, alpha=0.6, edgecolors='w', s=50, color=colors[color_idx])

                        # 添加对角线（理想线）
                        max_val = max(x_data.max(), y_data.max())
                        ax.plot([0, max_val], [0, max_val], 'r--', alpha=0.5, label='理想线')

                        # 计算回归线（使用自定义函数替代scipy.stats.linregress）
                        slope, intercept, r_value, r_squared = calculate_correlation_and_regression(x_data, y_data)

                        if slope is not None and intercept is not None:
                            x_line = np.array([0, max_val])
                            y_line = slope * x_line + intercept
                            # 使用R^2代替R²避免字体问题
                            ax.plot(x_line, y_line, 'g-', alpha=0.8, label=f'回归线 (R^2={r_squared:.3f})')

                        ax.set_xlabel(f"{standard_device} 日降雨量(mm)", fontsize=12)
                        ax.set_ylabel(f"{device} 日降雨量(mm)", fontsize=12)
                        ax.set_title(f"{device} 与 {standard_device} 对比", fontsize=14)
                        ax.grid(True, alpha=0.3)
                        ax.legend()

                # 隐藏多余的子图
                for idx in range(len(other_devices), rows * cols):
                    axes[idx].axis('off')

                plt.tight_layout()
                plt.savefig(os.path.join(output_dir, "与标准设备对比散点图.png"), dpi=150, bbox_inches='tight')
                plt.close()

                msg = f"✅ 与标准设备对比散点图已保存"
                results.append({"status": "success", "message": msg})

        # 4. 小时降雨强度对比图（最大降雨日）- 修改为显示所有设备
        plt.figure(figsize=(15, 10))

        # 修改：显示所有设备
        selected_devices = list(data_dict.keys())

        # 定义不同颜色和线型，增加颜色数量
        colors = ['blue', 'green', 'red', 'purple', 'orange', 'brown', 'pink', 'gray', 'cyan', 'magenta']
        line_styles = ['-', '--', '-.', ':', '-', '--', '-.', ':', '-', '--']

        for idx, device in enumerate(selected_devices):
            if len(data_dict[device]) > 0:
                # 找到最大降雨日
                max_rain_idx = data_dict[device]["日降雨量"].idxmax()
                max_rain_date = data_dict[device].loc[max_rain_idx, "日期"]

                # 提取小时数据
                hour_data = []
                for hour in range(24):
                    col = f"{hour}时"
                    if col in data_dict[device].columns:
                        hour_data.append(data_dict[device].loc[max_rain_idx, col])
                    else:
                        hour_data.append(np.nan)

                # 绘制小时降雨强度
                hours = list(range(24))
                color_idx = idx % len(colors)
                line_style_idx = idx % len(line_styles)

                plt.plot(hours, hour_data, linewidth=2, marker='o', markersize=4,
                         color=colors[color_idx], linestyle=line_styles[line_style_idx],
                         label=f"{device} ({max_rain_date.strftime('%Y-%m-%d')})")

        plt.title("各设备最大降雨日小时强度对比", fontsize=16, pad=20)
        plt.xlabel("小时", fontsize=14)
        plt.ylabel("小时降雨强度(mm)", fontsize=14)
        plt.xticks(range(0, 24))
        plt.grid(True, alpha=0.3)

        # 调整图例位置，避免重叠
        plt.legend(loc='upper left', bbox_to_anchor=(1.05, 1), borderaxespad=0., fontsize=10)
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, "小时降雨强度对比.png"), dpi=150, bbox_inches='tight')
        plt.close()

        msg = f"✅ 小时降雨强度对比图已保存"
        results.append({"status": "success", "message": msg})

        msg = f"✅ 基础可视化图表已保存到: {output_dir}"
        results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"⚠️ 生成可视化图表时出错: {str(e)}"
        results.append({"status": "warning", "message": msg})

    return results


def generate_analysis_report(data_dict, common_data, standard_device, output_dir, common_dates,
                             classification_results=None, accuracy_results=None):
    """
    生成分析报告
    """
    results = []

    try:
        report_path = os.path.join(output_dir, "降雨设备分析报告.txt")

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 70 + "\n")
            f.write("降雨设备计量准确度与雨量等级分析报告\n")
            f.write("=" * 70 + "\n\n")

            f.write(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"标准设备: {standard_device}\n")
            f.write(f"分析设备数量: {len(data_dict)}\n")
            f.write(f"设备列表: {', '.join(data_dict.keys())}\n\n")

            f.write("-" * 70 + "\n")
            f.write("1. 设备基本信息汇总\n")
            f.write("-" * 70 + "\n\n")

            # 设备基本信息
            for device in data_dict.keys():
                data = data_dict[device]
                total_rainfall = data["日降雨量"].sum()
                rainy_days = len(data)
                max_daily = data["日降雨量"].max()
                avg_daily = data["日降雨量"].mean()
                date_range = f"{data['日期'].min().strftime('%Y-%m-%d')} 至 {data['日期'].max().strftime('%Y-%m-%d')}"

                f.write(f"设备: {device}\n")
                f.write(f"  数据日期范围: {date_range}\n")
                f.write(f"  降雨天数: {rainy_days} 天\n")
                f.write(f"  累计降雨量: {total_rainfall:.2f} mm\n")
                f.write(f"  最大日降雨量: {max_daily:.2f} mm\n")
                f.write(f"  平均日降雨量: {avg_daily:.2f} mm\n\n")

            # 雨量等级分析结果
            if classification_results:
                f.write("-" * 70 + "\n")
                f.write("2. 雨量等级分类分析\n")
                f.write("-" * 70 + "\n\n")

                f.write("雨量等级划分标准（小时降雨量，单位：mm）:\n")
                f.write("  小雨: ≤2.5 mm\n")
                f.write("  中雨: 2.5~8 mm\n")
                f.write("  大雨: 8~16 mm\n")
                f.write("  暴雨: 16~20 mm\n")
                f.write("  短时强降雨: ≥20 mm\n\n")

                for device in data_dict.keys():
                    if device in classification_results:
                        results_dict = classification_results[device]
                        counts = results_dict["counts"]
                        rainfall = results_dict["rainfall"]
                        percentages = results_dict["percentages"]

                        f.write(f"设备: {device}\n")
                        f.write(f"  总观测小时数: {counts['总小时数']}\n")
                        f.write(f"  小雨: {counts['小雨']} 小时 ({percentages.get('小雨_占比(%)', 0)}%)\n")
                        f.write(f"  中雨: {counts['中雨']} 小时 ({percentages.get('中雨_占比(%)', 0)}%)\n")
                        f.write(f"  大雨: {counts['大雨']} 小时 ({percentages.get('大雨_占比(%)', 0)}%)\n")
                        f.write(f"  暴雨: {counts['暴雨']} 小时 ({percentages.get('暴雨_占比(%)', 0)}%)\n")
                        f.write(
                            f"  短时强降雨: {counts['短时强降雨']} 小时 ({percentages.get('短时强降雨_占比(%)', 0)}%)\n")
                        f.write(f"  总降雨量: {rainfall['总降雨量']:.2f} mm\n")
                        f.write(f"  小雨降雨量: {rainfall['小雨']:.2f} mm\n")
                        f.write(f"  中雨降雨量: {rainfall['中雨']:.2f} mm\n")
                        f.write(f"  大雨降雨量: {rainfall['大雨']:.2f} mm\n")
                        f.write(f"  暴雨降雨量: {rainfall['暴雨']:.2f} mm\n")
                        f.write(f"  短时强降雨量: {rainfall['短时强降雨']:.2f} mm\n\n")

                # 找出最强降雨设备
                max_torrential_device = None
                max_torrential_hours = -1

                max_storm_device = None
                max_storm_hours = -1

                max_heavy_device = None
                max_heavy_hours = -1

                for device in data_dict.keys():
                    if device in classification_results:
                        counts = classification_results[device]["counts"]

                        if counts["短时强降雨"] > max_torrential_hours:
                            max_torrential_hours = counts["短时强降雨"]
                            max_torrential_device = device

                        if counts["暴雨"] > max_storm_hours:
                            max_storm_hours = counts["暴雨"]
                            max_storm_device = device

                        if counts["大雨"] > max_heavy_hours:
                            max_heavy_hours = counts["大雨"]
                            max_heavy_device = device

                f.write("雨量等级特征分析:\n")
                if max_torrential_device and max_torrential_hours > 0:
                    f.write(f"  短时强降雨最多设备: {max_torrential_device} ({max_torrential_hours} 小时)\n")

                if max_storm_device and max_storm_hours > 0:
                    f.write(f"  暴雨最多设备: {max_storm_device} ({max_storm_hours} 小时)\n")

                if max_heavy_device and max_heavy_hours > 0:
                    f.write(f"  大雨最多设备: {max_heavy_device} ({max_heavy_hours} 小时)\n")

                f.write("\n")

            # 不同降雨等级下计量准确度分析结果
            if accuracy_results:
                f.write("-" * 70 + "\n")
                f.write("3. 不同降雨等级下计量准确度分析\n")
                f.write("-" * 70 + "\n\n")

                f.write(f"标准设备: {standard_device}\n\n")

                rainfall_classes = ["小雨", "中雨", "大雨", "暴雨", "短时强降雨"]

                # 各设备综合准确度评估
                f.write("3.1 各设备综合准确度评估:\n")
                for device in data_dict.keys():
                    if device != standard_device and device in accuracy_results:
                        # 计算加权平均准确度
                        total_samples = 0
                        weighted_mae = 0

                        for rain_class in rainfall_classes:
                            if rain_class in accuracy_results[device] and accuracy_results[device][rain_class][
                                "样本数"] > 0:
                                samples = accuracy_results[device][rain_class]["样本数"]
                                total_samples += samples
                                weighted_mae += accuracy_results[device][rain_class]["平均绝对误差(mm)"] * samples

                        if total_samples > 0:
                            overall_mae = weighted_mae / total_samples

                            # 评估等级
                            if overall_mae <= 0.5:
                                accuracy_level = "优秀"
                            elif overall_mae <= 1.0:
                                accuracy_level = "良好"
                            elif overall_mae <= 2.0:
                                accuracy_level = "一般"
                            else:
                                accuracy_level = "较差"

                            f.write(
                                f"  设备 {device}: 加权平均绝对误差 = {overall_mae:.3f} mm, 准确度等级: {accuracy_level}\n")

                f.write("\n")

                # 各降雨等级下准确度分析
                f.write("3.2 各降雨等级下准确度分析:\n")
                for rain_class in rainfall_classes:
                    f.write(f"\n  {rain_class}:\n")

                    # 收集该等级下所有设备的准确度
                    class_accuracies = []
                    for device in data_dict.keys():
                        if device != standard_device and device in accuracy_results:
                            if rain_class in accuracy_results[device] and accuracy_results[device][rain_class][
                                "样本数"] > 0:
                                result = accuracy_results[device][rain_class]
                                class_accuracies.append({
                                    "设备": device,
                                    "平均绝对误差": result["平均绝对误差(mm)"],
                                    "样本数": result["样本数"]
                                })

                    # 按平均绝对误差排序
                    if class_accuracies:
                        class_accuracies.sort(key=lambda x: x["平均绝对误差"])

                        best_device = class_accuracies[0]["设备"]
                        best_error = class_accuracies[0]["平均绝对误差"]
                        best_samples = class_accuracies[0]["样本数"]

                        worst_device = class_accuracies[-1]["设备"]
                        worst_error = class_accuracies[-1]["平均绝对误差"]
                        worst_samples = class_accuracies[-1]["样本数"]

                        f.write(
                            f"    最佳设备: {best_device} (平均绝对误差: {best_error:.3f} mm, 样本数: {best_samples})\n")
                        f.write(
                            f"    最差设备: {worst_device} (平均绝对误差: {worst_error:.3f} mm, 样本数: {worst_samples})\n")

                        # 输出所有设备
                        for acc in class_accuracies:
                            f.write(
                                f"    - {acc['设备']}: 平均绝对误差 = {acc['平均绝对误差']:.3f} mm (样本数: {acc['样本数']})\n")
                    else:
                        f.write(f"    无有效数据\n")

                f.write("\n")

                # 准确度趋势分析
                f.write("3.3 准确度趋势分析:\n")
                for device in data_dict.keys():
                    if device != standard_device and device in accuracy_results:
                        errors_by_class = []
                        for rain_class in rainfall_classes:
                            if rain_class in accuracy_results[device] and accuracy_results[device][rain_class][
                                "样本数"] > 0:
                                errors_by_class.append(accuracy_results[device][rain_class]["平均绝对误差(mm)"])
                            else:
                                errors_by_class.append(np.nan)

                        # 检查是否存在明显趋势
                        valid_errors = [err for err in errors_by_class if not np.isnan(err)]
                        if len(valid_errors) >= 3:
                            # 简单趋势判断：误差是否随降雨强度增加而增加
                            increasing = all(
                                valid_errors[i] <= valid_errors[i + 1] for i in range(len(valid_errors) - 1))
                            decreasing = all(
                                valid_errors[i] >= valid_errors[i + 1] for i in range(len(valid_errors) - 1))

                            if increasing and len(valid_errors) > 1:
                                f.write(f"  设备 {device}: 误差随降雨强度增加而增加\n")
                            elif decreasing and len(valid_errors) > 1:
                                f.write(f"  设备 {device}: 误差随降雨强度增加而减少\n")
                            else:
                                f.write(f"  设备 {device}: 误差与降雨强度无明显趋势关系\n")

                f.write("\n")

            f.write("-" * 70 + "\n")
            f.write("4. 设备间一致性分析\n")
            f.write("-" * 70 + "\n\n")

            if common_dates and len(common_dates) > 0:
                f.write(f"共有日期数量: {len(common_dates)} 天\n")
                f.write(f"共有日期范围: {common_dates[0]} 至 {common_dates[-1]}\n\n")

                # 计算与标准设备的一致性
                for device in data_dict.keys():
                    if device != standard_device and device in common_data:
                        # 获取共有日期的数据
                        merged_data = pd.merge(
                            common_data[standard_device][["日期", "日降雨量"]],
                            common_data[device][["日期", "日降雨量"]],
                            on="日期",
                            suffixes=('_标准', '_对比')
                        )

                        x_data = merged_data["日降雨量_标准"]
                        y_data = merged_data["日降雨量_对比"]

                        # 去除NaN值
                        mask = ~(np.isnan(x_data) | np.isnan(y_data))
                        if mask.sum() > 1:
                            x_clean = x_data[mask]
                            y_clean = y_data[mask]

                            # 计算统计指标
                            mean_diff = (y_clean - x_clean).mean()
                            mean_abs_diff = abs(y_clean - x_clean).mean()
                            max_diff = (y_clean - x_clean).abs().max()

                            # 相关系数（使用numpy.corrcoef）
                            if len(x_clean) > 1:
                                corr_matrix = np.corrcoef(x_clean, y_clean)
                                corr = corr_matrix[0, 1]

                                f.write(f"设备: {device} (与 {standard_device} 对比)\n")
                                f.write(f"  平均差异: {mean_diff:.2f} mm\n")
                                f.write(f"  平均绝对差异: {mean_abs_diff:.2f} mm\n")
                                f.write(f"  最大绝对差异: {max_diff:.2f} mm\n")
                                f.write(f"  相关系数: {corr:.4f}\n")

                                # 评估一致性
                                if corr > 0.95:
                                    f.write(f"  一致性评估: 优秀 (相关系数 > 0.95)\n")
                                elif corr > 0.9:
                                    f.write(f"  一致性评估: 良好 (0.9 < 相关系数 ≤ 0.95)\n")
                                elif corr > 0.8:
                                    f.write(f"  一致性评估: 一般 (0.8 < 相关系数 ≤ 0.9)\n")
                                else:
                                    f.write(f"  一致性评估: 较差 (相关系数 ≤ 0.8)\n")

                                f.write("\n")

            f.write("-" * 70 + "\n")
            f.write("5. 分析与建议\n")
            f.write("-" * 70 + "\n\n")

            # 检查异常设备
            if common_dates and len(common_dates) > 0:
                # 找出与标准设备差异最大的设备
                max_diff_device = None
                max_corr_device = None
                max_diff_value = -1
                max_corr_value = -1

                for device in data_dict.keys():
                    if device != standard_device and device in common_data:
                        merged_data = pd.merge(
                            common_data[standard_device][["日期", "日降雨量"]],
                            common_data[device][["日期", "日降雨量"]],
                            on="日期",
                            suffixes=('_标准', '_对比')
                        )

                        x_data = merged_data["日降雨量_标准"]
                        y_data = merged_data["日降雨量_对比"]

                        mask = ~(np.isnan(x_data) | np.isnan(y_data))
                        if mask.sum() > 1:
                            x_clean = x_data[mask]
                            y_clean = y_data[mask]

                            # 计算平均绝对差异
                            mean_abs_diff = abs(y_clean - x_clean).mean()

                            # 计算相关系数
                            if len(x_clean) > 1:
                                corr_matrix = np.corrcoef(x_clean, y_clean)
                                corr = corr_matrix[0, 1] if len(x_clean) > 1 else 0

                                if mean_abs_diff > max_diff_value:
                                    max_diff_value = mean_abs_diff
                                    max_diff_device = device

                                if corr > max_corr_value:
                                    max_corr_value = corr
                                    max_corr_device = device

                if max_diff_device:
                    f.write(f"与标准设备差异最大的设备: {max_diff_device} (平均绝对差异: {max_diff_value:.2f} mm)\n")
                    f.write("建议检查该设备的校准状态或安装位置。\n\n")

                if max_corr_device and max_corr_value > 0:
                    f.write(f"与标准设备一致性最好的设备: {max_corr_device} (相关系数: {max_corr_value:.4f})\n")
                    f.write("该设备的计量结果与标准设备高度一致。\n\n")

                # 检查是否存在系统偏差
                f.write("系统偏差检查:\n")
                for device in data_dict.keys():
                    if device != standard_device and device in common_data:
                        merged_data = pd.merge(
                            common_data[standard_device][["日期", "日降雨量"]],
                            common_data[device][["日期", "日降雨量"]],
                            on="日期",
                            suffixes=('_标准', '_对比')
                        )

                        x_data = merged_data["日降雨量_标准"]
                        y_data = merged_data["日降雨量_对比"]

                        mask = ~(np.isnan(x_data) | np.isnan(y_data))
                        if mask.sum() > 1:
                            x_clean = x_data[mask]
                            y_clean = y_data[mask]

                            # 计算平均偏差方向
                            mean_diff = (y_clean - x_clean).mean()

                            if abs(mean_diff) > 1.0:  # 如果平均偏差大于1mm
                                if mean_diff > 0:
                                    f.write(f"  - {device}: 系统偏高 {mean_diff:.2f} mm\n")
                                else:
                                    f.write(f"  - {device}: 系统偏低 {abs(mean_diff):.2f} mm\n")

            # 基于降雨等级准确度分析的建议
            if accuracy_results:
                f.write("\n基于降雨等级准确度分析的建议:\n")

                # 找出在不同降雨等级下表现差异大的设备
                for device in data_dict.keys():
                    if device != standard_device and device in accuracy_results:
                        errors = []
                        for rain_class in rainfall_classes:
                            if rain_class in accuracy_results[device] and accuracy_results[device][rain_class][
                                "样本数"] > 0:
                                errors.append(accuracy_results[device][rain_class]["平均绝对误差(mm)"])

                        if len(errors) >= 2:
                            error_range = max(errors) - min(errors)
                            if error_range > 1.0:
                                f.write(
                                    f"  - {device}: 在不同降雨等级下误差差异较大({error_range:.2f} mm)，建议检查其动态响应特性\n")

                # 针对短时强降雨的建议
                f.write("\n针对不同降雨等级的建议:\n")
                f.write("  - 小雨等级(<2.5mm): 关注设备的灵敏度和小雨检测能力\n")
                f.write("  - 中雨等级(2.5-8mm): 这是最常见的降雨强度，应确保所有设备在此范围内的准确性\n")
                f.write("  - 大雨等级(8-16mm): 关注设备的量程和排水能力\n")
                f.write("  - 暴雨等级(16-20mm): 检查设备的抗干扰能力和稳定性\n")
                f.write("  - 短时强降雨等级(>20mm): 关注设备的响应速度和峰值测量能力\n")

            f.write("\n6. 建议措施:\n")
            f.write("   - 定期校准所有雨量计设备\n")
            f.write("   - 检查设备安装位置是否规范\n")
            f.write("   - 清理设备漏斗，防止堵塞\n")
            f.write("   - 对差异较大的设备进行重点检查\n")
            f.write("   - 关注短时强降雨设备的性能表现\n")
            f.write("   - 根据雨量等级分析结果优化防洪预警策略\n")
            f.write("   - 针对不同降雨等级下的准确度差异，制定相应的校准和维护计划\n")

            f.write("\n" + "=" * 70 + "\n")
            f.write("报告结束\n")
            f.write("=" * 70 + "\n")

        msg = f"✅ 分析报告已保存到: {report_path}"
        results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"❌ 生成分析报告失败: {str(e)}"
        results.append({"status": "error", "message": msg})

    return results


def insert_images_to_excel(excel_path, image_dir):
    """
    将图表插入到Excel文件中
    """
    results = []

    try:
        wb = load_workbook(excel_path)

        # 创建新sheet用于存放图表
        if "分析图表" in wb.sheetnames:
            charts_sheet = wb["分析图表"]
            # 清空现有内容
            charts_sheet.delete_rows(1, charts_sheet.max_row)
        else:
            charts_sheet = wb.create_sheet(title="分析图表")

        # 设置图表工作表格式
        charts_sheet.sheet_view.showGridLines = False

        # 定义标题样式
        title_font = Font(name="微软雅黑", size=14, bold=True, color="366092")
        title_alignment = Alignment(horizontal="center", vertical="center")

        # 插入图表到Excel
        chart_files = [
            "累计降雨量对比.png",
            "日降雨量对比趋势.png",
            "与标准设备对比散点图.png",
            "小时降雨强度对比.png",
            "各设备雨量等级小时数对比.png",
            "各设备雨量等级小时数百分比对比.png",
            "各设备雨量等级累计降雨量对比.png",
            "各设备雨量等级小时数分布雷达图.png",
            "降雨等级准确度热力图.png",
            "降雨等级准确度对比图.png",
            "降雨等级准确度雷达图.png",
            "降雨等级样本分布图.png"
        ]

        # 修正行号问题：Excel行号从1开始
        # 每张图片占据25行（包括标题行和图片行）
        row_positions = [1, 27, 53, 79, 105, 131, 157, 183, 209, 235, 261, 287]  # 图片插入位置，从第1行开始

        current_row = 1

        for i, chart_file in enumerate(chart_files):
            chart_path = os.path.join(image_dir, chart_file)
            if os.path.exists(chart_path):
                try:
                    img = XLImage(chart_path)
                    # 调整图片大小
                    img.width = 600
                    img.height = 400

                    # 先添加标题
                    title_cell = f"A{current_row}"
                    charts_sheet[title_cell] = chart_file.replace(".png", "")
                    charts_sheet[title_cell].font = title_font
                    charts_sheet[title_cell].alignment = title_alignment

                    # 设置标题行高
                    charts_sheet.row_dimensions[current_row].height = 30

                    # 插入图片到标题下方
                    img_cell = f"A{current_row + 1}"
                    charts_sheet.add_image(img, img_cell)

                    current_row += 26  # 移动到下一张图片的位置（标题行+图片行+间隔）

                    msg = f"✅ 插入图表 {chart_file} 成功"
                    results.append({"status": "success", "message": msg})

                except Exception as e:
                    msg = f"⚠️ 插入图表 {chart_file} 失败: {str(e)}"
                    results.append({"status": "warning", "message": msg})
            else:
                msg = f"⚠️ 图表文件不存在: {chart_path}"
                results.append({"status": "warning", "message": msg})

        # 设置图表工作表的列宽
        charts_sheet.column_dimensions['A'].width = 100

        # 保存Excel文件
        wb.save(excel_path)
        msg = f"✅ 图表已插入到Excel文件中"
        results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"❌ 插入图表到Excel失败: {str(e)}"
        results.append({"status": "error", "message": msg})

    return results


def analyze_rainfall_devices(data_dict, standard_device, output_dir, progress_callback=None):
    """
    分析降雨设备数据 - 完整版
    """
    all_results = []

    try:
        # 检查标准设备是否存在
        if standard_device not in data_dict:
            msg = f"❌ 错误: 指定的标准设备 '{standard_device}' 不存在!"
            all_results.append({"status": "error", "message": msg})
            msg = f"可用的设备: {list(data_dict.keys())}"
            all_results.append({"status": "info", "message": msg})
            # 使用第一个设备作为默认标准
            standard_device = list(data_dict.keys())[0]
            msg = f"将使用 '{standard_device}' 作为标准设备"
            all_results.append({"status": "info", "message": msg})

        msg = f"\n📊 标准设备: {standard_device}"
        all_results.append({"status": "info", "message": msg})
        if progress_callback:
            progress_callback(msg)

        # 获取所有设备的共有日期
        common_dates = None
        for device, data in data_dict.items():
            device_dates = set(data["日期"].dt.date)
            if common_dates is None:
                common_dates = device_dates
            else:
                common_dates = common_dates.intersection(device_dates)

        common_dates = sorted(list(common_dates))
        msg = f"📅 共有日期数量: {len(common_dates)} 天"
        all_results.append({"status": "info", "message": msg})
        if progress_callback:
            progress_callback(msg)

        # 提取共有日期的数据
        common_data = {}
        for device, data in data_dict.items():
            data_copy = data.copy()
            data_copy["日期_日期"] = data_copy["日期"].dt.date
            common_mask = data_copy["日期_日期"].isin(common_dates)
            common_data[device] = data_copy[common_mask].copy()

        # 创建输出Excel文件
        output_path = os.path.join(output_dir, "降雨设备对比分析结果.xlsx")
        os.makedirs(output_dir, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 1. 汇总各设备日降雨量
            summary_data = []
            for device in data_dict.keys():
                data = data_dict[device]
                total_rainfall = data["日降雨量"].sum()
                rainy_days = len(data)
                max_daily = data["日降雨量"].max()
                avg_daily = data["日降雨量"].mean()

                summary_data.append({
                    "设备名称": device,
                    "累计降雨量(mm)": round(total_rainfall, 2),
                    "降雨天数": rainy_days,
                    "最大日降雨量(mm)": round(max_daily, 2),
                    "平均日降雨量(mm)": round(avg_daily, 2)
                })

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="设备汇总", index=False)

            # 2. 共有日期日降雨量对比
            if common_dates:
                comparison_data = {"日期": common_dates}
                for device in data_dict.keys():
                    if device in common_data:
                        device_rain = []
                        for date in common_dates:
                            mask = common_data[device]["日期_日期"] == date
                            if mask.any():
                                rain = common_data[device].loc[mask, "日降雨量"].values[0]
                                device_rain.append(rain)
                            else:
                                device_rain.append(np.nan)
                        comparison_data[device] = device_rain

                comparison_df = pd.DataFrame(comparison_data)
                comparison_df.to_excel(writer, sheet_name="日降雨量对比", index=False)

                # 计算与标准设备的差异
                if len(common_dates) > 0:
                    diff_data = {"日期": common_dates}
                    for device in data_dict.keys():
                        if device != standard_device and device in comparison_df.columns:
                            diff = comparison_df[device] - comparison_df[standard_device]
                            diff_data[f"{device}_差异"] = diff.round(2)
                            rel_diff = (diff / comparison_df[standard_device]) * 100
                            rel_diff = rel_diff.replace([np.inf, -np.inf], np.nan)
                            diff_data[f"{device}_相对差异%"] = rel_diff.round(2)

                    diff_df = pd.DataFrame(diff_data)
                    diff_df.to_excel(writer, sheet_name="与标准设备差异", index=False)

            msg = f"✅ 数据表格已生成"
            all_results.append({"status": "success", "message": msg})

        # 格式化Excel文件
        wb = load_workbook(output_path)

        # 格式化每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 读取工作表数据到DataFrame
            if ws.max_row > 1 and ws.max_column > 1:
                data = []
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                    data.append(row)

                # 第一行为表头
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])

                    # 清除工作表内容（保留格式设置）
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                        for cell in row:
                            cell.value = None

                    # 重新写入数据
                    for col_idx, col_name in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)

                    for row_idx, row_data in enumerate(df.values, 2):
                        for col_idx, cell_value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)

                    # 应用格式化
                    format_excel_worksheet(ws, df)

        # 保存工作簿
        wb.save(output_path)

        # 分析雨量等级分类
        msg = "\n📊 开始雨量等级分类分析..."
        all_results.append({"status": "info", "message": msg})
        if progress_callback:
            progress_callback(msg)

        classification_results, class_results = analyze_rainfall_classification(data_dict, output_dir)
        all_results.extend(class_results)

        # 分析不同降雨等级下的计量准确度
        msg = "\n📊 开始不同降雨等级下计量准确度分析..."
        all_results.append({"status": "info", "message": msg})
        if progress_callback:
            progress_callback(msg)

        accuracy_results, accuracy_analysis_results = analyze_accuracy_by_rainfall_class(data_dict, standard_device,
                                                                                         output_dir)
        all_results.extend(accuracy_analysis_results)

        # 生成雨量等级分类表格
        if classification_results:
            msg = "生成雨量等级分类表格..."
            all_results.append({"status": "info", "message": msg})
            if progress_callback:
                progress_callback(msg)

            class_table_results = generate_classification_tables(classification_results, output_path)
            all_results.extend(class_table_results)

        # 生成降雨等级准确度表格
        if accuracy_results:
            msg = "生成降雨等级准确度表格..."
            all_results.append({"status": "info", "message": msg})
            if progress_callback:
                progress_callback(msg)

            accuracy_table_results = generate_accuracy_tables(accuracy_results, output_path, standard_device)
            all_results.extend(accuracy_table_results)

        # 生成可视化图表
        msg = "生成可视化图表..."
        all_results.append({"status": "info", "message": msg})
        if progress_callback:
            progress_callback(msg)

        vis_results = generate_visualizations(data_dict, common_data, standard_device, output_dir, common_dates)
        all_results.extend(vis_results)

        # 生成雨量等级分类图表
        if classification_results:
            msg = "生成雨量等级分类图表..."
            all_results.append({"status": "info", "message": msg})
            if progress_callback:
                progress_callback(msg)

            class_chart_results = generate_classification_charts(classification_results, standard_device, output_dir)
            all_results.extend(class_chart_results)

        # 生成降雨等级准确度图表
        if accuracy_results:
            msg = "生成降雨等级准确度图表..."
            all_results.append({"status": "info", "message": msg})
            if progress_callback:
                progress_callback(msg)

            accuracy_chart_results = generate_accuracy_charts(accuracy_results, standard_device, output_dir)
            all_results.extend(accuracy_chart_results)

        # 生成分析报告
        msg = "生成分析报告..."
        all_results.append({"status": "info", "message": msg})
        if progress_callback:
            progress_callback(msg)

        report_results = generate_analysis_report(data_dict, common_data, standard_device, output_dir, common_dates,
                                                  classification_results, accuracy_results)
        all_results.extend(report_results)

        # 将图表插入Excel
        msg = "将图表插入Excel文件..."
        all_results.append({"status": "info", "message": msg})
        if progress_callback:
            progress_callback(msg)

        insert_results = insert_images_to_excel(output_path, output_dir)
        all_results.extend(insert_results)

        msg = f"\n✅ 分析结果已保存到: {output_path}"
        all_results.append({"status": "success", "message": msg})

    except Exception as e:
        msg = f"❌ 设备偏差分析失败: {str(e)}"
        all_results.append({"status": "error", "message": msg})
        if progress_callback:
            progress_callback(msg)

    return all_results


def run_complete_analysis(input_file, output_dir, standard_device, progress_callback=None):
    """
    完整分析入口函数
    """
    all_results = []

    msg = "=" * 70
    all_results.append({"status": "info", "message": msg})
    msg = "降雨设备计量准确度与雨量等级分析系统"
    all_results.append({"status": "info", "message": msg})
    msg = "=" * 70
    all_results.append({"status": "info", "message": msg})

    # 打印雨量等级划分标准
    msg = "\n📋 雨量等级划分标准（小时降雨量，单位：mm）:"
    all_results.append({"status": "info", "message": msg})
    msg = "  小雨: ≤2.5 mm"
    all_results.append({"status": "info", "message": msg})
    msg = "  中雨: 2.5~8 mm"
    all_results.append({"status": "info", "message": msg})
    msg = "  大雨: 8~16 mm"
    all_results.append({"status": "info", "message": msg})
    msg = "  暴雨: 16~20 mm"
    all_results.append({"status": "info", "message": msg})
    msg = "  短时强降雨: ≥20 mm"
    all_results.append({"status": "info", "message": msg})

    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        msg = f"❌ 错误: 输入文件不存在 - {input_file}"
        all_results.append({"status": "error", "message": msg})
        return all_results

    # 读取数据
    msg = "\n📖 正在读取数据..."
    all_results.append({"status": "info", "message": msg})
    if progress_callback:
        progress_callback(msg)

    data_dict, read_results = read_rainfall_data(input_file)
    all_results.extend(read_results)

    if not data_dict:
        msg = "❌ 错误: 未读取到任何有效数据!"
        all_results.append({"status": "error", "message": msg})
        return all_results

    msg = f"\n✅ 成功读取 {len(data_dict)} 个设备的数据"
    all_results.append({"status": "success", "message": msg})

    # 分析数据
    msg = "\n📊 正在分析数据..."
    all_results.append({"status": "info", "message": msg})
    if progress_callback:
        progress_callback(msg)

    analysis_results = analyze_rainfall_devices(data_dict, standard_device, output_dir, progress_callback)
    all_results.extend(analysis_results)

    msg = "\n" + "=" * 70
    all_results.append({"status": "info", "message": msg})
    msg = "分析完成!"
    all_results.append({"status": "success", "message": msg})
    msg = f"结果文件保存在: {output_dir}"
    all_results.append({"status": "info", "message": msg})
    msg = "主要输出内容:"
    all_results.append({"status": "info", "message": msg})
    msg = "  - Excel文件: 包含所有数据表格和图表"
    all_results.append({"status": "info", "message": msg})
    msg = "  - 分析报告: 详细的分析结果和建议"
    all_results.append({"status": "info", "message": msg})
    msg = "  - 图表文件: 多种可视化图表"
    all_results.append({"status": "info", "message": msg})
    msg = "  - 雨量等级分析: 小时降雨量等级分类统计"
    all_results.append({"status": "info", "message": msg})
    msg = "  - 降雨等级准确度分析: 不同降雨等级下各设备相对于标准设备的计量准确度"
    all_results.append({"status": "info", "message": msg})
    msg = "=" * 70
    all_results.append({"status": "info", "message": msg})

    return all_results


# 测试函数
if __name__ == "__main__":
    # 测试配置
    INPUT_FILE = "./降雨数据分析结果/批量降雨分析结果.xlsx"
    OUTPUT_DIR = "./降雨数据分析结果"
    STANDARD_DEVICE = "华云"

    # 运行完整分析
    results = run_complete_analysis(INPUT_FILE, OUTPUT_DIR, STANDARD_DEVICE)

    # 打印结果
    for result in results:
        print(result["message"])