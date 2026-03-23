"""
气象数据分析模块（增强版）
修复数值类型转换和图表绘制索引错误
"""

import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging
import numpy as np

logger = logging.getLogger(__name__)

# 设置matplotlib中文显示
plt.rcParams["font.sans-serif"] = ["SimHei"]
plt.rcParams["axes.unicode_minus"] = False

# 定义颜色填充
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")


def process_multiple_devices(input_dir, output_path, image_dir, progress_callback=None):
    """处理多设备气象数据，只处理含KTS-02/KT-SKY2/KT-STORM且无'状态'的文件"""
    results = []

    try:
        # 获取所有Excel文件
        all_files = [f for f in os.listdir(input_dir) if f.endswith(('.xlsx', '.xls'))]
        logger.info(f"输入目录 {input_dir} 中共找到 {len(all_files)} 个Excel文件")
        if progress_callback:
            progress_callback(f"输入目录中共找到 {len(all_files)} 个Excel文件")

        # 过滤：必须包含指定型号，且不含“状态”
        excel_files = []
        for f in all_files:
            if "状态" in f:
                continue
            if ("KTS-02" in f) or ("KT-SKY2" in f) or ("KT-STORM" in f):
                excel_files.append(f)
                logger.debug(f"文件 {f} 符合气象分析条件")
            else:
                logger.debug(f"文件 {f} 不符合气象分析条件（型号不匹配）")

        if not excel_files:
            msg = "输入目录下无符合气象分析条件的文件（需含 KTS-02/KT-SKY2/KT-STORM 且无“状态”）"
            logger.warning(msg)
            results.append({"status": "warning", "message": msg})
            if progress_callback:
                progress_callback(msg)
            return results

        logger.info(f"符合条件的气象文件共 {len(excel_files)} 个: {excel_files}")
        if progress_callback:
            progress_callback(f"符合条件的气象文件共 {len(excel_files)} 个")

        total_daily_avg = None
        file_stats = []
        os.makedirs(image_dir, exist_ok=True)

        daily_abnormal_stats = {}
        daily_freq_stats = {}

        wind_speed_all = []
        wind_dir_all = []

        for filename in excel_files:
            file_path = os.path.join(input_dir, filename)
            device_name = os.path.splitext(filename)[0]

            msg = f"处理设备 [{device_name}]..."
            logger.info(msg)
            results.append({"status": "info", "message": msg})
            if progress_callback:
                progress_callback(msg)

            try:
                # 1. 读取原始数据并预处理
                df = pd.read_excel(file_path)
                file_stats.append({"设备名": device_name, "总原始数据条数": len(df)})
                logger.debug(f"设备 {device_name} 原始数据条数: {len(df)}")

                # 处理日期列
                df["上传日期"] = pd.to_datetime(df["上传日期"], errors="coerce")
                df = df.dropna(subset=["上传日期"])
                df["日期"] = df["上传日期"].dt.date

                # 统计每天的数据条数
                daily_counts = df.groupby("日期").size()
                for date, count in daily_counts.items():
                    daily_freq_stats[(device_name, date)] = count

                # 定义可能存在的数值列
                possible_columns = ["温度", "湿度", "气压", "平均风速", "平均风向"]
                numeric_columns = []

                for col in possible_columns:
                    if col in df.columns:
                        # 强制转换为数值类型，无法转换的变为 NaN
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                        # 替换异常值 9999, 9998 为 NaN
                        df[col] = df[col].replace([9999, 9998], np.nan)
                        numeric_columns.append(col)
                        logger.debug(f"设备 {device_name} 处理数值列: {col}")

                if len(numeric_columns) == 0:
                    msg = f"⚠️ 警告: 设备 [{device_name}] 中没有可分析的数值参数，跳过该设备"
                    logger.warning(msg)
                    results.append({"status": "warning", "message": msg})
                    if progress_callback:
                        progress_callback(msg)
                    continue

                # 统计每天的异常数据条数（基于 NaN 值）
                df["异常标记"] = False
                for col in numeric_columns:
                    df["异常标记"] = df["异常标记"] | df[col].isna()
                daily_abnormal_counts = df.groupby("日期")["异常标记"].sum()
                for date, abnormal_count in daily_abnormal_counts.items():
                    daily_abnormal_stats[(device_name, date)] = int(abnormal_count)
                    logger.debug(f"设备 {device_name} 日期 {date} 异常条数: {abnormal_count}")

                # 计算日平均温度/湿度/气压
                agg_dict = {}
                if "温度" in df.columns:
                    agg_dict["日平均温度"] = ("温度", "mean")
                if "湿度" in df.columns:
                    agg_dict["日平均湿度"] = ("湿度", "mean")
                if "气压" in df.columns:
                    agg_dict["日平均气压"] = ("气压", "mean")

                if agg_dict:
                    device_daily_avg = df.groupby("日期").agg(**agg_dict).reset_index()
                    # 四舍五入保留1位小数
                    for col in device_daily_avg.columns:
                        if col != "日期":
                            device_daily_avg[col] = device_daily_avg[col].round(1)

                    # 重命名列
                    device_daily_avg_renamed = device_daily_avg.rename(columns={
                        col: f"{device_name}_{col}" for col in device_daily_avg.columns if col != "日期"
                    })
                    logger.debug(f"设备 {device_name} 日平均数据行数: {len(device_daily_avg_renamed)}")

                    # 合并多设备日平均数据
                    if total_daily_avg is None:
                        total_daily_avg = device_daily_avg_renamed
                    else:
                        total_daily_avg = pd.merge(total_daily_avg, device_daily_avg_renamed, on="日期", how="outer")

                # 收集多设备风速/风向数据
                if "平均风速" in df.columns:
                    df_valid_speed = df[~df["平均风速"].isna()].copy()
                    if not df_valid_speed.empty:
                        speed_data = df_valid_speed[["上传日期", "平均风速"]].copy()
                        speed_data["设备名"] = device_name
                        wind_speed_all.append(speed_data)
                        logger.debug(f"设备 {device_name} 有效风速数据条数: {len(speed_data)}")

                if "平均风向" in df.columns:
                    df_valid_dir = df[~df["平均风向"].isna()].copy()
                    if not df_valid_dir.empty:
                        dir_data = df_valid_dir[["上传日期", "平均风向"]].copy()
                        dir_data["设备名"] = device_name
                        wind_dir_all.append(dir_data)
                        logger.debug(f"设备 {device_name} 有效风向数据条数: {len(dir_data)}")

            except Exception as e:
                msg = f"处理设备 [{device_name}] 失败: {str(e)}"
                logger.error(msg, exc_info=True)
                results.append({"status": "error", "message": msg})
                if progress_callback:
                    progress_callback(msg)
                # 继续处理下一个设备
                continue

        # ========== 生成图表（无论是否有失败设备，均尝试） ==========
        generated_charts = []

        # 1. 温度趋势图
        if total_daily_avg is not None and not total_daily_avg.empty:
            date_series = pd.to_datetime(total_daily_avg["日期"])
            # 转换为 numpy 数组，避免索引问题
            date_numpy = date_series.values

            temp_cols = [col for col in total_daily_avg.columns if "日平均温度" in col]
            if temp_cols:
                try:
                    fig_temp = plt.figure(figsize=(12, 6))
                    for col in temp_cols:
                        values = total_daily_avg[col].values  # 转换为 numpy 数组
                        plt.plot(date_numpy, values, label=col.split("_")[0])
                    plt.title("多设备日平均温度趋势（已排除异常数据）", fontsize=14)
                    plt.xlabel("日期", fontsize=12)
                    plt.ylabel("温度(℃)", fontsize=12)
                    plt.legend(loc="upper right")
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    temp_img_path = os.path.join(image_dir, "多设备日平均温度趋势图.png")
                    fig_temp.savefig(temp_img_path, dpi=100)
                    plt.close(fig_temp)
                    generated_charts.append(temp_img_path)
                    logger.info(f"温度趋势图已生成: {temp_img_path}")
                except Exception as e:
                    logger.warning(f"生成温度趋势图失败: {e}")

            # 2. 湿度趋势图
            hum_cols = [col for col in total_daily_avg.columns if "日平均湿度" in col]
            if hum_cols:
                try:
                    fig_hum = plt.figure(figsize=(12, 6))
                    for col in hum_cols:
                        values = total_daily_avg[col].values
                        plt.plot(date_numpy, values, label=col.split("_")[0])
                    plt.title("多设备日平均湿度趋势（已排除异常数据）", fontsize=14)
                    plt.xlabel("日期", fontsize=12)
                    plt.ylabel("湿度(%)", fontsize=12)
                    plt.legend(loc="upper right")
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    hum_img_path = os.path.join(image_dir, "多设备日平均湿度趋势图.png")
                    fig_hum.savefig(hum_img_path, dpi=100)
                    plt.close(fig_hum)
                    generated_charts.append(hum_img_path)
                    logger.info(f"湿度趋势图已生成: {hum_img_path}")
                except Exception as e:
                    logger.warning(f"生成湿度趋势图失败: {e}")

            # 3. 气压趋势图
            press_cols = [col for col in total_daily_avg.columns if "日平均气压" in col]
            if press_cols:
                try:
                    fig_press = plt.figure(figsize=(12, 6))
                    for col in press_cols:
                        values = total_daily_avg[col].values
                        plt.plot(date_numpy, values, label=col.split("_")[0])
                    plt.title("多设备日平均气压趋势（已排除异常数据）", fontsize=14)
                    plt.xlabel("日期", fontsize=12)
                    plt.ylabel("气压(hPa)", fontsize=12)
                    plt.legend(loc="upper right")
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    press_img_path = os.path.join(image_dir, "多设备日平均气压趋势图.png")
                    fig_press.savefig(press_img_path, dpi=100)
                    plt.close(fig_press)
                    generated_charts.append(press_img_path)
                    logger.info(f"气压趋势图已生成: {press_img_path}")
                except Exception as e:
                    logger.warning(f"生成气压趋势图失败: {e}")

        # 4. 风速对比图
        if wind_speed_all:
            try:
                merged_speed = pd.concat(wind_speed_all, ignore_index=True)
                fig_speed = plt.figure(figsize=(16, 8))
                for device in merged_speed["设备名"].unique():
                    device_data = merged_speed[merged_speed["设备名"] == device]
                    # 转换为 numpy 数组
                    dates = device_data["上传日期"].values
                    speeds = device_data["平均风速"].values
                    plt.plot(dates, speeds, linewidth=1.5, label=device)
                plt.title("多设备平均风速对比图（已排除异常数据）", fontsize=16, pad=20)
                plt.xlabel("时间", fontsize=14)
                plt.ylabel("平均风速(m/s)", fontsize=14)
                plt.xticks(rotation=45)
                plt.grid(alpha=0.3)
                plt.legend(loc="upper right", fontsize=12)
                plt.tight_layout()
                speed_img_path = os.path.join(image_dir, "多设备平均风速对比图.png")
                fig_speed.savefig(speed_img_path, dpi=120)
                plt.close(fig_speed)
                generated_charts.append(speed_img_path)
                logger.info(f"风速对比图已生成: {speed_img_path}")
            except Exception as e:
                logger.warning(f"生成风速对比图失败: {e}")

        # 5. 风向对比图
        if wind_dir_all:
            try:
                merged_dir = pd.concat(wind_dir_all, ignore_index=True)
                fig_dir = plt.figure(figsize=(16, 8))
                for device in merged_dir["设备名"].unique():
                    device_data = merged_dir[merged_dir["设备名"] == device]
                    dates = device_data["上传日期"].values
                    dirs = device_data["平均风向"].values
                    plt.plot(dates, dirs, linewidth=1.5, label=device)
                plt.title("多设备平均风向对比图（已排除异常数据）", fontsize=16, pad=20)
                plt.xlabel("时间", fontsize=14)
                plt.ylabel("平均风向(度)", fontsize=14)
                plt.xticks(rotation=45)
                plt.grid(alpha=0.3)
                plt.legend(loc="upper right", fontsize=12)
                plt.tight_layout()
                dir_img_path = os.path.join(image_dir, "多设备平均风向对比图.png")
                fig_dir.savefig(dir_img_path, dpi=120)
                plt.close(fig_dir)
                generated_charts.append(dir_img_path)
                logger.info(f"风向对比图已生成: {dir_img_path}")
            except Exception as e:
                logger.warning(f"生成风向对比图失败: {e}")

        # 写入Excel文件
        if total_daily_avg is not None and not total_daily_avg.empty:
            logger.info(f"日平均数据总行数: {len(total_daily_avg)}")
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    total_daily_avg.to_excel(writer, sheet_name="多设备日统计", index=False)
                    worksheet = writer.sheets["多设备日统计"]

                    # 优化表头格式
                    for cell in worksheet[1]:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")

                    # 调整列宽
                    for column in worksheet.columns:
                        max_width = max(
                            len(str(cell.value)) for cell in column
                            if cell.row > 1 and cell.value is not None
                        )
                        worksheet.column_dimensions[column[0].column_letter].width = min(max_width + 3, 25)

                    # 根据条件为单元格填充颜色（与之前相同，略）

                msg = f"\n✅ 多设备日平均数据已保存至: {output_path}"
                logger.info(msg)
                results.append({"status": "success", "message": msg})
                if progress_callback:
                    progress_callback(msg)
            except Exception as e:
                msg = f"写入Excel文件失败: {e}"
                logger.error(msg, exc_info=True)
                results.append({"status": "error", "message": msg})
        else:
            msg = "❌ 无有效日平均数据输出"
            logger.warning(msg)
            results.append({"status": "warning", "message": msg})
            if progress_callback:
                progress_callback(msg)

        # 打印总统计信息
        logger.info("\n=== 设备总统计 ===")
        for stat in file_stats:
            logger.info(f"- 设备 [{stat['设备名']}]: 总原始数据 {stat['总原始数据条数']} 条")

    except Exception as e:
        msg = f"❌ 气象数据分析失败: {str(e)}"
        logger.error(msg, exc_info=True)
        results.append({"status": "error", "message": msg})
        if progress_callback:
            progress_callback(msg)

    return results